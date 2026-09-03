#!/usr/bin/env python3
"""
SkillPort Admin > Reports > Templates > "Asset Activity by User" scraper.

This is a separate script from skillport_scraper.py - that one scrapes the Library (the student
course catalog), this one scrapes the result of one specific template from the Admin console's
Reports section, paging through every page (Page 1 of 32, etc.) and collecting all rows into one
Excel file.

The proven login/driver code is imported from skillport_scraper.py (the same functions that are
already working against this exact site) - no need to rewrite it.

Install:  pip install selenium openpyxl
Run:
  set SKILLPORT_PASS=SkillTalha$26
  python asset_activity_report_scraper.py
  python asset_activity_report_scraper.py --headless
  python asset_activity_report_scraper.py --max-pages 3      (for testing, only 3 pages)

IMPORTANT - do NOT pass --headless the first time you run this script. Leave the browser open and
watch that every step clicks the right place or not (Group filter, Display Options, Preview).
SkillPort's admin panel is an old JSP/frame-based app whose exact HTML I never got the chance to
see live - which is why the selectors are mostly based on VISIBLE TEXT / LABEL (more reliable when
the exact element ID isn't known), but if some step fails the script will stop on its own and give
you the chance to complete that step manually (this won't happen in headless mode).
"""

import argparse, json, os, re, subprocess, sys, time
from pathlib import Path

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select

# ---- Reuse the already-working login/driver machinery from the existing scraper ----
from skillport_scraper import (
    BASE, USERNAME, PASSWORD, WAIT, PAGE_SETTLE,
    log, norm, make_driver, force_kill_driver, login, js_click, find_any, by_text,
    page_ready, shot, OUT_DIR,
)

# ===================== CONFIG =====================
# Copied directly from the address bar the user showed - if SkillPort ever changes this
# template's internal ID, re-open it manually once (Reports > Templates > User > Learning
# Activity > double-click "Asset Activity by User") and paste the new URL here.
TEMPLATE_EDIT_URL = (
    f"{BASE}/admin/edittemplate.action?templateId=27&templateType=Default"
    f"&reportKey=learning_activity_asset_activity_by_user&callingScreen=templateList"
    f"&originalTemplateId=27"
    f"&templateDescription=Detailed%20asset%20activity%20data%2C%20sorted%20by%20user."
    f"&originalRootTemplateId=27"
)

# Fallback only - used when the SuperAdmin Settings > Report Scraper page (SkillportScraperSettings
# table, read live via _fetch_group_name_from_db below) has never been configured or the database
# isn't reachable from wherever this script happens to be running. Same search term is used for
# both the search box and the exact group selected - real groups here are single codes (e.g.
# "LC_17") where the two were never meaningfully different in practice.
DEFAULT_GROUP_NAME = "LC_17"

# How long to wait for the report itself to render after Preview. An "All Groups" report across a
# year is thousands of pages and can sit on "loading" for many minutes; the old 90s ceiling gave up
# and scraped whatever was on screen. Overridable with --report-load-timeout.
REPORT_LOAD_TIMEOUT = 900

# How long to wait for each SUBSEQUENT page's grid after clicking Next. Was a hardcoded 30s that
# carried on regardless when it lapsed - which ended a 1302-page run at page 21. Overridable with
# --page-load-timeout.
PAGE_LOAD_TIMEOUT = 180

# How many times to relaunch the browser and carry on after it dies mid-run. Each restart costs a
# login, a report regeneration and a jump back to the last page - minutes, against the hours of
# scraping it saves. Generous because the failure it recovers from (Chrome killed under memory
# pressure during a multi-hour run) can genuinely happen several times in one report.
MAX_BROWSER_RESTARTS = 20

OUTPUT_XLSX = OUT_DIR / "asset_activity_by_user.xlsx"
OUTPUT_TXT = OUT_DIR / "asset_activity_by_user.txt"
OUTPUT_CSV_CHECKPOINT = OUT_DIR / "asset_activity_by_user_checkpoint.csv"
# Records which page was last completed, so --resume can pick up there instead of re-scraping from
# page 1. A 3000-page run is hours long; losing it to a dropped connection is not acceptable.
OUTPUT_PROGRESS_JSON = OUT_DIR / "asset_activity_by_user_progress.json"

# Same database the .NET API's own DefaultConnection points at (src/API/appsettings.json) - the
# SuperAdmin Settings > Report Scraper page writes to this exact table, so reading it here is what
# makes "whichever group name I add in the UI, that's the one that runs" actually true. Env-var
# overridable, hardcoded fallback matching this project's existing convention for this script (see
# SKILLPORT_USER/SKILLPORT_PASS in skillport_scraper.py) rather than a missing-config hard failure.
DB_SERVER = os.environ.get("SKILLSETS_DB_SERVER", "144.76.246.72")
DB_NAME = os.environ.get("SKILLSETS_DB_NAME", "SoftSkillSetLocal")
DB_USER = os.environ.get("SKILLSETS_DB_USER", "WORKWELL")
DB_PASSWORD = os.environ.get("SKILLSETS_DB_PASSWORD", "A12dadf125@@@")
# ==================================================


def _fetch_group_name_from_db():
    """Reads the current Group Name from dbo.SkillportScraperSettings (the SuperAdmin Settings >
    Report Scraper page) via sqlcmd, so a group changed in the UI takes effect on this script's very
    next run with no code edit. Falls back to DEFAULT_GROUP_NAME - not a hard failure - if sqlcmd
    isn't on PATH, the database is unreachable, or no SuperAdmin has ever saved a group yet, since a
    single missing/unreachable setting shouldn't block scraping the group that was already working."""
    try:
        result = subprocess.run(
            [
                "sqlcmd", "-S", DB_SERVER, "-U", DB_USER, "-P", DB_PASSWORD, "-d", DB_NAME, "-C",
                "-h", "-1", "-W",
                "-Q", "SET NOCOUNT ON; SELECT TOP 1 GroupName FROM dbo.SkillportScraperSettings ORDER BY SkillportScraperSettingsId DESC;",
            ],
            capture_output=True, text=True, timeout=15,
        )
        lines = [line.strip() for line in result.stdout.splitlines() if line.strip() and not line.strip().startswith("(")]
        if result.returncode == 0 and lines and lines[0].lower() != "null":
            log(f"Got the Group Name from Settings: '{lines[0]}'")
            return lines[0]
    except Exception as e:
        log(f"WARNING: Could not get the Group Name from Settings ({e}) - using default '{DEFAULT_GROUP_NAME}'")
        return DEFAULT_GROUP_NAME

    log(f"WARNING: Settings table is empty or sqlcmd failed - using default '{DEFAULT_GROUP_NAME}'")
    return DEFAULT_GROUP_NAME


# ===================== FRAME-AWARE HELPERS =====================
# SkillPort's admin console is a classic multi-frame JSP app - the content we need may sit inside
# an <iframe>/<frame> rather than the top-level document, and possibly inside a frame nested
# inside ANOTHER frame. These helpers search every frame RECURSIVELY (depth-first, bounded depth)
# instead of assuming top-level or only one level of nesting.

def _iter_contexts(driver, _max_depth=3):
    """Yields (name, switch_fn) for the top document plus every frame, RECURSIVELY - a frame can
    itself contain another frame (this admin console's template editor turned out to be nested
    two levels deep for at least the Preview button, which a one-level-only scan never reached).
    Depth-first, bounded to _max_depth levels as a guard against pathological frame nesting."""

    def make_switch(path):
        def switch():
            driver.switch_to.default_content()
            for i in path:
                frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
                driver.switch_to.frame(frames[i])
        return switch

    def walk(path, depth):
        name = "top" if not path else f"frame{list(path)}"
        switch = make_switch(path)
        yield name, switch
        if depth >= _max_depth:
            return
        try:
            switch()
            frame_count = len(driver.find_elements(By.CSS_SELECTOR, "iframe, frame"))
        except Exception:
            frame_count = 0
        for i in range(frame_count):
            yield from walk(path + (i,), depth + 1)

    yield from walk((), 0)


def find_in_any_context(driver, locators, timeout=10):
    """Like find_any(), but also searches inside every frame. Leaves the driver switched into
    whichever context the element was found in (caller can keep acting there)."""
    end = time.time() + timeout
    while time.time() < end:
        for name, switch in _iter_contexts(driver):
            try:
                switch()
            except Exception:
                continue
            el = find_any(driver, locators, timeout=0.3)
            if el:
                return el
        time.sleep(0.4)
    driver.switch_to.default_content()
    return None


def goto_url_any_context(driver, url):
    driver.switch_to.default_content()
    driver.get(url)
    page_ready(driver)
    time.sleep(PAGE_SETTLE)


# ===================== STEP 1: open the template editor =====================

def open_asset_activity_template(driver):
    log("Opening the Asset Activity by User template edit page (direct URL)")
    goto_url_any_context(driver, TEMPLATE_EDIT_URL)
    el = find_in_any_context(driver, [by_text("*", "groups / users"), by_text("*", "activity dates")], WAIT)
    if not el:
        shot(driver, "template_page_not_found")
        raise RuntimeError(
            "The template edit page did not open, or the expected content was not found. "
            "Check the screenshot output/debug_template_page_not_found_*.png.")
    log("Template edit page opened")


def _dump_debug_html(driver, anchor_element, filename):
    """Saves the outerHTML of a small ancestor container around anchor_element, so a failed
    selector guess can be pinpointed from the real markup instead of another round of blind
    guessing from a screenshot alone. Pass anchor_element=None when the lookup found nothing at all
    to anchor on - the whole document is dumped instead, which is exactly the case where there is no
    other way to see what the page actually rendered."""
    try:
        html = driver.execute_script(
            "let el = arguments[0]; "
            "if (!el) return document.body ? document.body.outerHTML : ''; "
            "for (let i = 0; i < 5 && el.parentElement; i++) el = el.parentElement; "
            "return el.outerHTML;", anchor_element)
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        (OUT_DIR / filename).write_text(html or "", encoding="utf-8")
        log(f"Debug HTML saved: output/{filename}")
    except Exception as e:
        log(f"Debug HTML dump fail: {e}")


def _expand_section(driver, header_text, count_selector):
    """This template edit page has collapsible accordion sections (Groups / Users, Display
    Options) that start CLOSED - their fields don't exist in a usable/visible state until the
    header is clicked. Detects collapsed-vs-expanded by counting matching elements before/after
    clicking, and clicks again if the first click collapsed it instead of opening it."""
    header = find_in_any_context(driver, [by_text("*", header_text)], 8)
    if not header:
        return None
    before = driver.find_elements(By.CSS_SELECTOR, count_selector)
    js_click(driver, header)
    time.sleep(0.8)
    after = driver.find_elements(By.CSS_SELECTOR, count_selector)
    if len(after) <= len(before):
        js_click(driver, header)
        time.sleep(0.8)
    return header


# ===================== STEP 2: Groups / Users filter =====================

def is_all_groups(group_name):
    """Whether the configured group means "every group" rather than one specific org code. Accepts
    the handful of spellings someone might reasonably type into Settings > Report Scraper."""
    return (group_name or "").strip().lower() in {"all", "all groups", "allgroups", "*"}


def configure_group_filter(driver, group_name):
    # Local names, not module-level constants - shadows every GROUP_SEARCH_TERM/GROUP_TO_SELECT
    # reference below with whatever group main() resolved for this run (Settings table or default),
    # without having to touch each of those call sites individually.
    GROUP_SEARCH_TERM = GROUP_TO_SELECT = group_name

    # "All Groups" is the template's own default in the destination box, and every check further
    # down this function exists to make sure that default was REPLACED by a specific group. When the
    # whole point is to report on everything, that default is the correct end state - so confirm it
    # is there and stop, rather than searching for a group named "All".
    if is_all_groups(group_name):
        log("Group filter: ALL GROUPS - leaving the template's own 'All Groups' selection in place")
        _expand_section(driver, "groups / users", "input[type='text']")
        if not find_in_any_context(driver, [by_text("*", "all groups")], 8):
            shot(driver, "all_groups_not_present")
            raise RuntimeError(
                "'All Groups' was not found in the destination box - the template's default has changed. "
                "Check output/debug_all_groups_not_present_*.png.")
        log("Group filter: 'All Groups' confirmed")
        return

    log(f"Group filter: search '{GROUP_SEARCH_TERM}', select '{GROUP_TO_SELECT}'")

    # This section (an accordion panel, separate from the "Groups / Users" tab above it) starts
    # collapsed - the search box/tree don't exist yet until this header is clicked. This is
    # exactly what was missing before ("the 'Search' button was not found" - it genuinely wasn't in
    # the DOM in a usable state yet).
    section_header = _expand_section(driver, "groups / users", "input[type='text']")
    if not section_header:
        shot(driver, "groups_users_section_not_found")
        raise RuntimeError("The 'Groups / Users' collapsible section header was not found.")

    # Make sure the "Groups" radio (not "Users") is selected. This is the one reliable, unambiguous
    # anchor we have in this section - everything else below is found by walking forward through
    # the DOM from this exact element, instead of matching on text like "Search" that may not be
    # unique on the page (there could be another "Search" elsewhere, e.g. a nav search box).
    radio = find_in_any_context(driver, [
        (By.XPATH, "//input[@type='radio' and (contains(@value,'roup') or contains(@id,'roup'))]"),
    ], 8)
    if not radio:
        shot(driver, "groups_radio_not_found")
        raise RuntimeError("The 'Groups' radio button was not found.")
    if not radio.is_selected():
        js_click(driver, radio)
        time.sleep(0.5)

    # Walk forward from the radio in document order to the next visible plain text input - that's
    # the group search box, wherever exactly it sits in the markup.
    search_box = driver.execute_script("""
        const anchor = arguments[0];
        const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_ELEMENT);
        let passedAnchor = false;
        while (walker.nextNode()) {
            const node = walker.currentNode;
            if (node === anchor) { passedAnchor = true; continue; }
            if (!passedAnchor) continue;
            if (node.tagName === 'INPUT' && (node.type === 'text' || !node.getAttribute('type'))
                && node.offsetParent !== null) {
                return node;
            }
        }
        return null;
    """, radio)

    if not search_box:
        _dump_debug_html(driver, radio, "debug_groups_section_html.txt")
        shot(driver, "group_search_box_not_found")
        raise RuntimeError(
            "The group search box (after the Groups radio) was not found. "
            "Send output/debug_groups_section_html.txt so the exact structure can be checked.")

    search_box.clear()
    search_box.send_keys(GROUP_SEARCH_TERM)
    time.sleep(0.3)
    actual_value = (search_box.get_attribute("value") or "").strip()
    if actual_value != GROUP_SEARCH_TERM:
        driver.execute_script(
            "arguments[0].value = arguments[1];"
            "arguments[0].dispatchEvent(new Event('input', {bubbles: true}));"
            "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
            search_box, GROUP_SEARCH_TERM)
        actual_value = (search_box.get_attribute("value") or "").strip()
    log(f"Typed into the search box - its value is now: '{actual_value}'")
    if actual_value != GROUP_SEARCH_TERM:
        shot(driver, "search_box_value_mismatch")
        raise RuntimeError(
            f"Could not type '{GROUP_SEARCH_TERM}' into the search box (value: '{actual_value}'). "
            "Check the screenshot output/debug_search_box_value_mismatch_*.png.")

    # Walk forward from search_box (not a bare page-wide "search" text match, which could hit an
    # unrelated element elsewhere on the page) to the next element whose own visible text is
    # exactly "search".
    search_btn = driver.execute_script("""
        const anchor = arguments[0];
        const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_ELEMENT);
        let passedAnchor = false;
        while (walker.nextNode()) {
            const node = walker.currentNode;
            if (node === anchor) { passedAnchor = true; continue; }
            if (!passedAnchor) continue;
            const text = (node.innerText || node.value || '').trim().toLowerCase();
            if (text === 'search' && node.offsetParent !== null) {
                return node;
            }
        }
        return null;
    """, search_box)
    if not search_btn:
        _dump_debug_html(driver, search_box, "debug_groups_section_html.txt")
        shot(driver, "search_button_not_found")
        raise RuntimeError(
            "The 'Search' button was not found. "
            "Send output/debug_groups_section_html.txt so the exact structure can be checked.")

    js_click(driver, search_btn)
    time.sleep(PAGE_SETTLE)

    # Resolve straight to the ROW ELEMENT the grid's SelectionModel listens on (class contains
    # "x-grid3-row"), rather than finding "some element with this text" and hoping afterwards that
    # it happens to sit inside a row. Clicking the innermost cell text does NOT register as a real
    # selection here (the row highlights from the click's own focus styling, but SelectionModel
    # never fires - proven by the Add button's "enabled" twin staying hidden), so the row element is
    # what we actually need.
    #
    # Searching cell-first then walking up used to fail outright for a SHORT group code: with a
    # value like "ss" the page-wide text fallback matched something that was never in a grid at all,
    # the walk-up found no row ancestor, and the run died with the group unset. Collecting every
    # exact-text match and keeping only those that genuinely resolve to a row removes that whole
    # class of failure - a candidate that isn't in a row simply isn't a candidate.
    find_row_js = """
        const target = arguments[0].trim().toLowerCase();

        function rowOf(el) {
            let e = el;
            while (e && !(typeof e.className === 'string' && e.className.includes('x-grid3-row'))) {
                e = e.parentElement;
            }
            return e;
        }

        const rows = [];
        for (const el of document.querySelectorAll('*')) {
            if (el.children.length !== 0) continue;
            const text = (el.innerText || el.textContent || '').trim().toLowerCase();
            if (text !== target) continue;
            const row = rowOf(el);
            if (row && !rows.includes(row)) rows.push(row);
        }
        if (!rows.length) return null;

        // Prefer a row inside the grid that also renders an "Org Code" header - that is the
        // left-hand SEARCH RESULTS grid. The right-hand destination box shows the very same group
        // text once an earlier run already added it, and selecting that one instead leaves "Add"
        // with nothing to move, silently reporting on every group instead of just this one.
        function inResultsGrid(row) {
            let g = row;
            for (let i = 0; i < 10 && g.parentElement; i++) {
                g = g.parentElement;
                if ((g.innerText || '').toLowerCase().includes('org code')) return true;
            }
            return false;
        }
        return rows.find(inResultsGrid) || rows[0];
    """

    # Search EVERY frame, not just whichever one the driver happens to be sitting in - this console
    # is frame-based throughout, and an element is only clickable while its own frame is the active
    # context, so the frame that found the row is the frame we must stay in to click it.
    row_el = None
    for name, switch in _iter_contexts(driver):
        try:
            switch()
            row_el = driver.execute_script(find_row_js, GROUP_TO_SELECT)
        except Exception:
            continue
        if row_el:
            log(f"  [group filter] found the '{GROUP_TO_SELECT}' row (context={name})")
            break

    if not row_el:
        _dump_debug_html(driver, None, "debug_group_row_html.txt")
        shot(driver, "group_row_not_found")
        raise RuntimeError(
            f"Could not find the search-result grid row for '{GROUP_TO_SELECT}' (the search did run). "
            "Send output/debug_group_row_html.txt and output/debug_group_row_not_found_*.png.")

    row_el.click()
    time.sleep(0.5)

    # This template renders TWO overlapping "Add" controls sharing the same position - a static
    # grey arrow (id "disabledUGCopyButton", always visible, does nothing when clicked) and a real
    # one (id "enabledUGCopyButton", hidden until a row is genuinely selected in the grid's
    # SelectionModel). Clicking the always-visible grey one is why the group filter previously
    # silently no-opped and the report ran against every group instead of just this one.
    enabled_add = find_in_any_context(driver, [(By.ID, "enabledUGCopyButton")], 6)
    if not enabled_add or enabled_add.value_of_css_property("display") == "none":
        shot(driver, "add_arrow_not_enabled")
        raise RuntimeError(
            "The 'Add' button never entered its enabled state - the row selection did not register. "
            "Check output/debug_add_arrow_not_enabled_*.png.")
    clickable = driver.execute_script(
        "return arguments[0].querySelector('img, a, table') || arguments[0];", enabled_add)
    clickable.click()
    time.sleep(0.8)

    # Verify the destination box now shows ONLY the searched group, not the template's own default
    # "All Groups" entry (which otherwise silently reports on every group instead of just this one -
    # confirmed live: the destination box starts with "All Groups" pre-selected on this template,
    # and a real "Add" replaces it, but a no-op "Add" leaves it in place with no visible error).
    still_has_all_groups = find_in_any_context(driver, [by_text("*", "all groups")], 3)
    if still_has_all_groups:
        shot(driver, "all_groups_still_selected")
        raise RuntimeError(
            "'All Groups' is still present in the destination box even after 'Add' - the group filter "
            f"was not restricted to '{GROUP_TO_SELECT}'. Check output/debug_all_groups_still_selected_*.png.")

    dest_has_group = find_in_any_context(driver, [by_text("*", GROUP_TO_SELECT)], 4)
    if not dest_has_group:
        shot(driver, "group_add_not_confirmed")
        raise RuntimeError(
            f"'{GROUP_TO_SELECT}' could not be confirmed in the destination box even after clicking 'Add'. "
            "Check output/debug_group_add_not_confirmed_*.png.")
    log("Group filter set (All Groups removed, only the target group selected)")


# ===================== STEP 2b: Activity Dates =====================

def configure_activity_dates(driver, amount=1, unit="Year"):
    """Sets the report's Activity Dates to "Previous <amount> <unit>(s)".

    The template defaults to one month, which silently truncates the history - anything older simply
    is not in the export, and nothing about the output says so. Widening it here is what makes a
    full year's activity actually reach the import.

    Controls are identified by what they CONTAIN rather than by fixed ids or DOM position - the ids
    are not stable across templates, and the panel's layout shifts depending on which tab is open.
    """
    log(f"Activity Dates: setting Previous {amount} {unit}(s)")

    _expand_section(driver, "activity dates", "input[type='radio']")

    radio = find_in_any_context(driver, [
        (By.XPATH, "//*[contains(translate(normalize-space(.),'PREVIOUS','previous'),'previous')]"
                   "/preceding::input[@type='radio'][1]"),
        (By.XPATH, "//input[@type='radio' and (contains(@id,'revious') or contains(@value,'revious'))]"),
    ], 8)
    if radio and not radio.is_selected():
        js_click(driver, radio)
        time.sleep(0.4)
    elif not radio:
        # Not fatal on its own: on some templates "Previous" is already the active choice and has no
        # separately addressable radio. The real test is whether the period dropdown below gets set.
        log("  (Activity Dates' 'Previous' radio was not found - probably already selected, continuing anyway)")

    # Identify the unit dropdown by WHAT IT CONTAINS - the one offering Day/Week/Month/Year - rather
    # than by position relative to the radio. Walking forward from the radio picked up the User
    # Status dropdown ("Activated"/"Deactivated") that sits further down the same panel, because the
    # period selector is not necessarily the first <select> after it in document order.
    #
    # The amount box is then the nearest visible text input BEFORE that dropdown, which is exactly
    # how the control reads on screen: [Previous] [1] [Year(s)].
    # The period control is NOT a native <select>. Confirmed live: every <select> on the page is
    # something else entirely (User Status, timezone, Multiple/Single Row, completion status) - the
    # Day/Week/Month/Year picker is an ExtJS combo, i.e. a readonly <input> showing the current value
    # with its list rendered elsewhere in the DOM and only populated when opened.
    #
    # So it is located by the VALUE it is currently displaying ("Month(s)", "Year(s)", ...), which is
    # the one thing that reliably identifies it, and the amount box is the visible text input just
    # before it - exactly how the row reads on screen: [Previous] [1] [Month(s)].
    set_dates_js = """
        const amount = String(arguments[0]);
        const periodPattern = /^(day|week|month|year)\\(s\\)$/i;

        const visible = (el) => el && el.offsetParent !== null;
        const inputs = [...document.querySelectorAll('input')].filter(visible);

        const combo = inputs.find(i => periodPattern.test((i.value || '').trim()));
        if (!combo) {
            return {
                ok: false,
                reason: 'period combo not found',
                inputValuesSeen: inputs.map(i => (i.value || '').trim()).filter(v => v).slice(0, 15),
            };
        }

        const before = inputs.filter(i =>
            i !== combo && (i.type === 'text' || !i.type) &&
            (combo.compareDocumentPosition(i) & Node.DOCUMENT_POSITION_PRECEDING));
        const box = before.length ? before[before.length - 1] : null;
        if (!box) return { ok: false, reason: 'amount box not found', combo: combo.value };

        box.value = amount;
        box.dispatchEvent(new Event('input', { bubbles: true }));
        box.dispatchEvent(new Event('change', { bubbles: true }));

        return { ok: true, amount: box.value, current: (combo.value || '').trim() };
    """

    # Search EVERY frame - this console is frame-based, and the Activity Dates panel does not have to
    # live in whichever frame the radio above was found in.
    applied = None
    for name, switch in _iter_contexts(driver):
        try:
            switch()
            result = driver.execute_script(set_dates_js, amount, unit)
        except Exception:
            continue
        if result and result.get("ok"):
            applied = result
            log(f"  [activity dates] context={name}")
            break
        if result and applied is None:
            applied = result  # keep the most informative failure for the error message

    if not applied or not applied.get("ok"):
        shot(driver, "activity_dates_not_set")
        _dump_debug_html(driver, None, "debug_activity_dates.txt")
        raise RuntimeError(
            f"Could not set the Activity Dates: {applied}. "
            "Check output/debug_activity_dates_not_set_*.png and output/debug_activity_dates.txt.")

    # Now the unit itself. An ExtJS combo only commits a value through its own click handling, so
    # this drives it the way a person does: click to open the list, then click the matching item.
    # Setting .value directly would show the right text and submit the old value.
    wanted = f"{unit}(s)".lower()
    if applied.get("current", "").lower() == wanted:
        log(f"Activity Dates set: Previous {applied['amount']} {applied['current']} (unit already correct)")
        return

    combo = driver.execute_script("""
        const periodPattern = /^(day|week|month|year)\\(s\\)$/i;
        return [...document.querySelectorAll('input')]
            .filter(i => i.offsetParent !== null)
            .find(i => periodPattern.test((i.value || '').trim())) || null;
    """)
    if combo is None:
        shot(driver, "period_combo_lost")
        raise RuntimeError("Could not find the period combo again. Check output/debug_period_combo_lost_*.png.")

    js_click(driver, combo)
    time.sleep(0.6)

    # The opened list is rendered outside the combo (ExtJS appends it to the document), so it is
    # matched on its visible text anywhere on the page rather than inside the combo's own subtree.
    picked = driver.execute_script("""
        const wanted = arguments[0];
        const items = [...document.querySelectorAll('div,li,td,span')].filter(el =>
            el.children.length === 0 &&
            (el.innerText || '').trim().toLowerCase() === wanted &&
            el.offsetParent !== null);
        if (!items.length) {
            return { ok: false, seen: [...document.querySelectorAll('div,li,td,span')]
                .filter(el => el.children.length === 0 && el.offsetParent !== null)
                .map(el => (el.innerText || '').trim())
                .filter(t => /\\(s\\)$/.test(t)).slice(0, 12) };
        }
        items[0].click();
        return { ok: true };
    """, wanted)

    time.sleep(0.5)

    if not picked or not picked.get("ok"):
        shot(driver, "period_option_not_found")
        raise RuntimeError(
            f"The '{unit}(s)' option was not found in the dropdown: {picked}. "
            "Check output/debug_period_option_not_found_*.png.")

    final = driver.execute_script("""
        const periodPattern = /^(day|week|month|year)\\(s\\)$/i;
        const el = [...document.querySelectorAll('input')]
            .filter(i => i.offsetParent !== null)
            .find(i => periodPattern.test((i.value || '').trim()));
        return el ? (el.value || '').trim() : null;
    """)

    if not final or final.lower() != wanted:
        shot(driver, "period_not_applied")
        raise RuntimeError(
            f"After selecting unit '{unit}(s)', the combo is still showing '{final}'. "
            "Check output/debug_period_not_applied_*.png.")

    log(f"Activity Dates set: Previous {applied['amount']} {final}")


# ---- Explicit From/To range (used for --date-from/--date-to, i.e. the "Custom" mode) ----
# Written out rather than taken from strftime("%b"): strftime is locale-dependent, and on a machine
# with a non-English locale it would produce month names the site never uses.
MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _date_renderings(value):
    """One dd/mm/yyyy input rendered into every format this site might be displaying dates in.

    Confirmed live: this template's Activity Dates boxes show dates as "2026-Aug-01", not as
    "01/09/2026" - which is why matching only slash dates would find no date fields at all. Rather
    than hardcode that one format (another template, or a different SkillPort locale setting, could
    easily use another), every plausible rendering is computed here and the browser side picks
    whichever one matches what the fields are ALREADY showing.
    """
    from datetime import datetime
    d = datetime.strptime(value.strip(), "%d/%m/%Y")
    return {
        "ymd_abbr": f"{d.year:04d}-{MONTH_ABBR[d.month - 1]}-{d.day:02d}",   # 2026-Sep-01
        "ymd_dash": f"{d.year:04d}-{d.month:02d}-{d.day:02d}",               # 2026-09-01
        "dmy_slash": f"{d.day:02d}/{d.month:02d}/{d.year:04d}",              # 01/09/2026
        "mdy_slash": f"{d.month:02d}/{d.day:02d}/{d.year:04d}",              # 09/01/2026
    }


# Finds the two date boxes, switches the panel onto its date-range option, and writes both values in
# the format the boxes are already using. Kept as one script so all of it runs against the same
# frame in a single round trip - the fields, the radio that governs them and the ExtJS component
# that owns them all live together, and splitting it risked acting on a panel that had re-rendered
# in between.
SET_DATE_RANGE_JS = r"""
    const fromByFormat = arguments[0], toByFormat = arguments[1];

    const visible = (el) => el && el.offsetParent !== null;

    // Recognised display formats, most specific first.
    const FORMATS = [
        ['ymd_abbr', /^\d{4}-[A-Za-z]{3}-\d{1,2}$/],
        ['ymd_dash', /^\d{4}-\d{1,2}-\d{1,2}$/],
        ['slash',    /^\d{1,2}\/\d{1,2}\/\d{4}$/],
    ];

    function formatOf(value) {
        for (const [name, re] of FORMATS) if (re.test(value)) return name;
        return null;
    }

    const allInputs = [...document.querySelectorAll('input')];
    const fields = allInputs.filter(el =>
        visible(el) &&
        (!el.type || el.type === 'text') &&
        formatOf((el.value || '').trim()));

    if (fields.length < 2) {
        return {
            ok: false,
            reason: 'need two date fields, found ' + fields.length,
            inputValuesSeen: allInputs.filter(visible)
                .map(i => (i.value || '').trim()).filter(v => v).slice(0, 20),
        };
    }

    const fromField = fields[0], toField = fields[1];
    let fmt = formatOf((fromField.value || '').trim());

    // dd/mm and mm/dd look identical whenever both numbers are 12 or under, so decide from the
    // sample itself where possible: a component above 12 can only be a day. Falls back to dd/mm,
    // which is what this deployment uses elsewhere, and the caller verifies the result either way.
    if (fmt === 'slash') {
        const parts = [fromField, toField].map(el => (el.value || '').split('/'));
        const firstOver12 = parts.some(p => parseInt(p[0], 10) > 12);
        const secondOver12 = parts.some(p => parseInt(p[1], 10) > 12);
        fmt = secondOver12 && !firstOver12 ? 'mdy_slash' : 'dmy_slash';
    }

    const wantedFrom = fromByFormat[fmt], wantedTo = toByFormat[fmt];
    if (!wantedFrom || !wantedTo) {
        return { ok: false, reason: 'unsupported date format: ' + fmt };
    }

    // The panel offers "Previous [N] [Month(s)]" and a date-range row as alternatives, and only the
    // selected one is actually applied. The radio governing the range is the last radio appearing
    // BEFORE the first date box in document order - a positional anchor, because the visible label
    // wording ("Range", "Between", "From") is not something this template could be relied on to use.
    const radios = [...document.querySelectorAll("input[type='radio']")].filter(visible);
    const preceding = radios.filter(r =>
        fromField.compareDocumentPosition(r) & Node.DOCUMENT_POSITION_PRECEDING);
    const rangeRadio = preceding.length ? preceding[preceding.length - 1] : null;
    let radioClicked = false;
    if (rangeRadio && !rangeRadio.checked) {
        rangeRadio.click();
        radioClicked = true;
    }

    function write(el, text) {
        // ExtJS keeps its own internal value alongside the DOM one. Writing only the DOM value
        // makes the box LOOK right while the form still submits the old date, so go through the
        // component when there is one. (This console is ExtJS 3 - the input's own id is the
        // component id; newer versions suffix it with -inputEl, so both are tried.)
        try {
            if (window.Ext && Ext.getCmp && el.id) {
                const cmp = Ext.getCmp(el.id) || Ext.getCmp(String(el.id).replace(/-inputEl$/, ''));
                if (cmp && cmp.setValue) {
                    cmp.setValue(text);
                    return;
                }
            }
        } catch (e) { /* fall through to the plain DOM write below */ }

        const setter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype, 'value').set;
        const wasReadOnly = el.readOnly;
        el.readOnly = false;                 // date pickers often mark their input readonly
        setter.call(el, text);
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        el.dispatchEvent(new Event('blur', { bubbles: true }));
        el.readOnly = wasReadOnly;
    }

    write(fromField, wantedFrom);
    write(toField, wantedTo);

    return {
        ok: true,
        format: fmt,
        radioClicked: radioClicked,
        radioChecked: rangeRadio ? !!rangeRadio.checked : null,
        expectedFrom: wantedFrom,
        expectedTo: wantedTo,
        from: (fromField.value || '').trim(),
        to: (toField.value || '').trim(),
    };
"""


# Re-reads the two date boxes after the dust settles. A separate pass on purpose: selecting the
# range radio can re-render the panel, and an ExtJS field that rejects a value reverts silently, so
# what the boxes read a moment later is the only trustworthy confirmation.
READ_DATE_FIELDS_JS = r"""
    const visible = (el) => el && el.offsetParent !== null;
    const FORMATS = [
        /^\d{4}-[A-Za-z]{3}-\d{1,2}$/,
        /^\d{4}-\d{1,2}-\d{1,2}$/,
        /^\d{1,2}\/\d{1,2}\/\d{4}$/,
    ];
    const fields = [...document.querySelectorAll('input')].filter(el =>
        visible(el) && (!el.type || el.type === 'text') &&
        FORMATS.some(re => re.test((el.value || '').trim())));
    if (fields.length < 2) return null;
    return { from: (fields[0].value || '').trim(), to: (fields[1].value || '').trim() };
"""


def configure_activity_dates_range(driver, date_from, date_to):
    """Sets the report's Activity Dates to an explicit FROM/TO range (the "Custom" mode).

    Dates are given to this function as dd/mm/yyyy. What gets TYPED into the page is whatever format
    the boxes are already displaying, detected at run time - this template shows "2026-Aug-01", and
    writing a slash date into it would either be rejected or, worse, silently reinterpreted.

    Controls are identified by what they CONTAIN or by position relative to the date boxes rather
    than by fixed ids - the ids are not stable across templates, and the panel's layout shifts
    depending on which tab is open.
    """
    log(f"Activity Dates: setting the range {date_from} to {date_to}")

    _expand_section(driver, "activity dates", "input[type='radio']")

    from_renderings = _date_renderings(date_from)
    to_renderings = _date_renderings(date_to)

    # Search EVERY frame - this console is frame-based, and the Activity Dates panel does not have
    # to live in whichever frame the previous step was working in.
    applied = None
    for name, switch in _iter_contexts(driver):
        try:
            switch()
            result = driver.execute_script(SET_DATE_RANGE_JS, from_renderings, to_renderings)
        except Exception:
            continue
        if result and result.get("ok"):
            applied = result
            log(f"  [activity dates] context={name}, detected date format: {result['format']}")
            break
        if result and applied is None:
            applied = result  # keep the most informative failure for the error message

    if not applied or not applied.get("ok"):
        shot(driver, "activity_dates_not_set")
        _dump_debug_html(driver, None, "debug_activity_dates.txt")
        raise RuntimeError(
            f"Could not set the Activity Dates range: {applied}. "
            "Check output/debug_activity_dates_not_set_*.png and output/debug_activity_dates.txt.")

    if applied.get("radioClicked"):
        log("  the date-range radio was selected (it was on 'Previous N ...' before)")

    # Let the panel settle, then read the boxes back rather than trusting the values returned above:
    # selecting the range radio can re-render the row, and a rejected value reverts without any
    # visible error. Getting this wrong exports a different window than the one asked for, which is
    # invisible in the finished file.
    #
    # Deliberately searches EVERY frame again from scratch rather than jumping straight back to
    # whichever context the write above happened in: _iter_contexts recomputes frame paths by
    # POSITION each time it's called, and the write above can itself trigger an ExtJS re-render that
    # changes how many frames exist or reorders them - so a "context=frame[2]" from the write is not
    # guaranteed to still mean the same physical frame a moment later. Confirmed live: this exact
    # mismatch was why an earlier version raised "Activity Dates range not confirmed" even though a
    # screenshot taken at the same moment showed the boxes already reading correctly.
    time.sleep(0.8)
    confirmed = None
    for name, switch in _iter_contexts(driver):
        try:
            switch()
            result = driver.execute_script(READ_DATE_FIELDS_JS)
        except Exception:
            continue
        if result is None:
            continue
        if result.get("from") == applied["expectedFrom"] and result.get("to") == applied["expectedTo"]:
            confirmed = result
            break
        if confirmed is None:
            confirmed = result  # keep the first non-null reading in case no context matches exactly

    final_from = (confirmed or applied).get("from")
    final_to = (confirmed or applied).get("to")
    expected_from = applied["expectedFrom"]
    expected_to = applied["expectedTo"]

    if final_from != expected_from or final_to != expected_to:
        shot(driver, "activity_dates_not_applied")
        _dump_debug_html(driver, None, "debug_activity_dates.txt")
        raise RuntimeError(
            f"The date boxes read '{final_from}' to '{final_to}' after being set to "
            f"'{expected_from}' to '{expected_to}' (detected format: {applied['format']}) - the page "
            "did not accept the values. Check output/debug_activity_dates_not_applied_*.png and "
            "output/debug_activity_dates.txt.")

    log(f"Activity Dates set: {final_from} to {final_to} "
        f"(from {date_from} to {date_to}, written as {applied['format']})")


# ===================== STEP 3: Display Options - select everything =====================

def expand_and_select_all_display_options(driver):
    log("Expanding Display Options and selecting all fields")

    # The real field-selection checkboxes live on the "Filter Options" TAB, not the "Groups / Users"
    # tab this function used to be called right after - confirmed live: without switching tabs first,
    # this searched whatever happened to be on screen and found only 2 unrelated checkboxes
    # (reportToDateCB/filterToDateCB from Activity Dates), never the real ~55 Display Fields ones.
    tab = find_in_any_context(driver, [by_text("*", "filter options")], 8)
    if not tab:
        shot(driver, "filter_options_tab_not_found")
        raise RuntimeError("The 'Filter Options' tab was not found.")
    js_click(driver, tab)
    time.sleep(1.5)

    # This specific accordion panel does not respond to a click on its header bar/text (unlike
    # "Groups / Users", which does) - confirmed live: only its small collapse/expand icon
    # (class "x-tool x-tool-toggle") actually toggles the panel. Only click it if still collapsed;
    # clicking twice would re-collapse it.
    header_div = driver.execute_script("""
        return [...document.querySelectorAll('div.x-panel-header')].find(el =>
            (el.innerText||'').trim().toLowerCase() === 'display options');
    """)
    if not header_div:
        shot(driver, "display_options_header_not_found")
        raise RuntimeError("The 'Display Options' header was not found.")

    is_collapsed = "x-panel-collapsed" in driver.execute_script(
        "return arguments[0].parentElement.className;", header_div)
    if is_collapsed:
        toggle_icon = driver.execute_script(
            "return arguments[0].querySelector('.x-tool-toggle');", header_div)
        if not toggle_icon:
            shot(driver, "display_options_toggle_not_found")
            raise RuntimeError("The 'Display Options' expand/collapse icon was not found.")
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", header_div)
        time.sleep(0.3)
        ActionChains(driver).move_to_element(toggle_icon).pause(0.2).click().perform()
        time.sleep(1.5)

    still_collapsed = "x-panel-collapsed" in driver.execute_script(
        "return arguments[0].parentElement.className;", header_div)
    if still_collapsed:
        shot(driver, "display_options_still_collapsed")
        raise RuntimeError("'Display Options' still did not expand even after clicking the toggle icon.")

    checked, skipped_disabled = 0, 0
    for cb in driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']"):
        try:
            if not cb.is_displayed():
                continue
            if not cb.is_enabled():
                skipped_disabled += 1
                continue
            if not cb.is_selected():
                js_click(driver, cb)
                checked += 1
        except Exception:
            continue

    # This template has ~55 display-field checkboxes once genuinely expanded - a low count here is
    # the exact symptom of interacting with the wrong (collapsed/wrong-tab) region again, so fail
    # loudly instead of silently producing a report missing most of its columns.
    total_visible = len(driver.execute_script("""
        return [...document.querySelectorAll('input[type=checkbox]')].filter(cb => {
            const r = cb.getBoundingClientRect();
            return r.width > 0 && r.height > 0;
        });
    """))
    if total_visible < 20:
        shot(driver, "display_options_suspiciously_few_checkboxes")
        raise RuntimeError(
            f"Only {total_visible} checkboxes were visible (expected ~55) - probably the wrong "
            "section got expanded. Check output/debug_display_options_suspiciously_few_checkboxes_*.png.")

    log(f"Display Options: checked {checked} new checkbox(es), skipped {skipped_disabled} already-locked/disabled "
        f"one(s), {total_visible} total visible")


# ===================== STEP 4: Preview =====================

def click_preview(driver, allow_existing_report=False):
    """allow_existing_report is for the resume path. Skillport restores the previous report session
    on login, so a resumed run can land straight in the report VIEWER - where there is no Preview
    button at all, because the report it would generate is already on screen. Treating that as a
    fatal error threw away a resume that was otherwise perfectly positioned to continue.

    Deliberately opt-in: on a FRESH run an already-rendered report is a stale one from an earlier
    session, generated with different settings, and scraping it would silently export the wrong
    data. There the missing button stays an error."""
    if allow_existing_report and any_context_has_valid_table(driver):
        log("Report has already rendered (resume) - no need to click Preview")
        return

    log("Clicking Preview")
    btn = find_in_any_context(driver, [
        by_text("*", "preview"),
        (By.CSS_SELECTOR, "input[type='button'][value*='preview' i], input[type='submit'][value*='preview' i]"),
        (By.CSS_SELECTOR, "img[alt*='preview' i], img[title*='preview' i], a[title*='preview' i]"),
    ], 15)
    if not btn:
        shot(driver, "preview_button_not_found")
        # Dump what's actually visible in every context so a repeat failure is diagnosable
        # straight from the log, without another round trip.
        for name, switch in _iter_contexts(driver):
            try:
                switch()
                texts = [norm(e.text) for e in driver.find_elements(By.CSS_SELECTOR, "a, button, input, span, div")
                         if e.is_displayed() and norm(e.text)]
                log(f"  [preview debug] context={name}: {len(texts)} visible text elements, "
                    f"sample: {texts[:15]}")
            except Exception as e:
                log(f"  [preview debug] context={name}: error reading ({e})")
        driver.switch_to.default_content()
        raise RuntimeError("The 'Preview' button/link was not found.")
    js_click(driver, btn)
    time.sleep(PAGE_SETTLE)

    # The report can take a while to actually render ("report is loading please wait") - a fixed
    # short sleep here was the real bug: scraping started before the real data table existed yet,
    # so it fell back to whatever small table WAS there (the pagination bar), producing exactly
    # "Page of 32" as the only scraped content. Poll for a table that actually validates against
    # real column headers instead of guessing a fixed wait time.
    # An "All Groups" report over a year is a very different job from one group over a month - it can
    # legitimately sit on "report is loading" for many minutes before the first page appears. The old
    # 90-second ceiling gave up long before that and then carried on regardless, which is how an
    # empty/wrong table got scraped. Waits far longer now, and says so while it waits, so a long load
    # is visibly progress rather than a hang.
    log(f"Waiting for the report to load (max {REPORT_LOAD_TIMEOUT // 60} min)...")
    end = time.time() + REPORT_LOAD_TIMEOUT
    started = time.time()
    found = False
    next_heartbeat = 30
    while time.time() < end:
        if any_context_has_valid_table(driver):
            found = True
            break
        waited = int(time.time() - started)
        if waited >= next_heartbeat:
            log(f"  ... still loading ({waited // 60}m {waited % 60}s)")
            next_heartbeat += 30
        time.sleep(2.0)

    if not found:
        shot(driver, "report_never_loaded")
        raise RuntimeError(
            f"No results table was found even after waiting {REPORT_LOAD_TIMEOUT // 60} minutes. "
            "This used to start scraping anyway and pick up a wrong/empty table - now it stops instead. "
            "Check output/debug_report_never_loaded_*.png, or increase --report-load-timeout.")

    log(f"Report loaded (in {int(time.time() - started)}s), found the real data table")
    time.sleep(1.0)


# ===================== STEP 5: scrape the results table across every page =====================

PAGE_LABEL_RE = re.compile(r"Page\s+(\d+)\s+of\s+(\d+)", re.I)


# Confirmed live from the user's own screenshot of the real header row - used to VALIDATE a
# candidate table is actually the results grid, not the pagination bar's own <table> (which also
# has rows/cells - "Page", an input, "of", "722" - and could win a plain "most rows" comparison on
# some pages, which is exactly what was happening: counts/page-numbers instead of real data).
EXPECTED_HEADER_HINTS = [
    "user id", "first name", "last name", "asset title", "asset id", "asset type",
    "group name", "display first name", "display last name",
]


# A Skillport login username as it appears in the grid's per-person group band - either the newer
# "<digits>LC<digits>" account style (e.g. "10lc387010") or an older plain login (e.g. "adachhn").
# Deliberately narrow: one bare token, no spaces, so a stray one-cell row holding a sentence, a
# count, or a date is never mistaken for an identity.
GROUP_BAND_RE = re.compile(r"^[A-Za-z0-9._@-]{3,100}$")


def _direct_rows(table):
    """DIRECT child <tr> of table (or its <tbody>), NOT every descendant. This admin console's
    report viewer renders every single cell's text inside its OWN one-cell nested <table> (a
    per-cell text-formatting wrapper, confirmed live via outerHTML) - table.find_elements(By.TAG_NAME,
    "tr") descends into THOSE too and returns their inner one-cell <tr> as if it were a real row of
    its own, which is what produced the "First Name" / "Last Name" / ... rows stacked one-per-row in
    column A instead of a normal grid. Direct children only, via XPath, sidesteps this entirely."""
    # Excludes the report viewer's own column-width spacer row (confirmed live:
    # <tr height="0" aria-hidden="true"> with one empty <td> per column, purely for layout) - a
    # precise marker rather than a cell-count heuristic, since that spacer row otherwise has a
    # normal-looking cell count (one per real column).
    xpath = "./tbody/tr[not(@aria-hidden='true')]"
    rows = table.find_elements(By.XPATH, xpath)
    return rows or table.find_elements(By.XPATH, "./tr[not(@aria-hidden='true')]")


def _direct_cells(row):
    """DIRECT child <td>/<th> of row - same reasoning as _direct_rows: a real row's own cells,
    not the inner <tr> of some other cell's per-cell text-wrapper table three levels down."""
    return row.find_elements(By.XPATH, "./td|./th")


def _looks_like_results_table(table):
    try:
        for row in _direct_rows(table)[:5]:
            cells = _direct_cells(row)
            if len(cells) < 10:
                continue
            header_text = " | ".join(norm(c.text).lower() for c in cells)
            if sum(1 for hint in EXPECTED_HEADER_HINTS if hint in header_text) >= 2:
                return True
    except Exception:
        return False
    return False


def _table_candidates_in_current_context(driver):
    """(row_count, table) for every table in whichever context the driver is CURRENTLY switched
    into, sorted by row count descending. Does not touch frame switching itself. Scored by DIRECT
    child row count (see _direct_rows) - the report viewer's own outer wrapper tables have a huge
    RECURSIVE tr count (every per-cell text-wrapper table's own <tr> included) that used to win this
    comparison over the real, much-smaller-by-recursive-count data grid table."""
    tables = driver.find_elements(By.TAG_NAME, "table")
    scored = []
    for t in tables:
        try:
            n = len(_direct_rows(t))
        except Exception:
            n = 0
        scored.append((n, t))
    scored.sort(key=lambda x: x[0], reverse=True)
    return scored


def any_context_has_valid_table(driver):
    """True if a real, header-validated results table exists in ANY frame right now - used to
    poll "has the report actually finished loading yet" without caring which frame it lands in."""
    for name, switch in _iter_contexts(driver):
        try:
            switch()
        except Exception:
            continue
        if any(_looks_like_results_table(t) for t in driver.find_elements(By.TAG_NAME, "table")):
            return True
    driver.switch_to.default_content()
    return False


def _find_results_table(driver, strict=False):
    """Picks the results grid <table>, searching EVERY FRAME (not just whichever frame the driver
    happens to be in) - this admin console is frame-based, and the results grid after Preview can
    live in a different frame than the one active when Preview was clicked. Confirmed live: with
    only the top-level document searched, every candidate found was pagination/toolbar junk
    ("Page of 32", "Edit Save Template...") even though the real data table was clearly visible on
    screen - it was sitting in a frame that was never checked. Within each frame, ranks candidates
    by row count but VALIDATES each against known real column headers before accepting it.

    strict=True disables the largest-candidate fallback below and returns None instead. Use it once
    paging is under way: mid-run, "no validated table" means the next page simply hasn't rendered
    yet, and the fallback would hand back whatever junk happens to be on screen (live: a 3-row,
    blank-header toolbar table in the top context on page 21 of 1302), which then extracts as empty
    and looks exactly like a clean end of data. The caller waits and retries instead."""
    for name, switch in _iter_contexts(driver):
        try:
            switch()
        except Exception:
            continue
        for n, t in _table_candidates_in_current_context(driver):
            if n > 1 and _looks_like_results_table(t):
                try:
                    header_cells = _direct_cells(_direct_rows(t)[0])
                    snippet = " | ".join(norm(c.text) for c in header_cells[:6])
                except Exception:
                    snippet = "(header read fail)"
                log(f"  [table check] SELECTED (context={name}): {n} rows, headers: {snippet}")
                return t

    if strict:
        return None

    # Nothing validated anywhere - gather candidates from every context, log each one's header
    # snippet (with which context it came from) so a future failure is diagnosable straight from
    # this log, then fall back to the single largest candidate across all of them.
    log("WARNING: no table matched the expected headers (User ID / First Name / Asset Title...) "
        "in any frame. Headers of all candidate tables:")
    all_candidates = []
    for name, switch in _iter_contexts(driver):
        try:
            switch()
        except Exception:
            continue
        for n, t in _table_candidates_in_current_context(driver)[:5]:
            try:
                header_cells = _direct_cells(_direct_rows(t)[0])
                snippet = " | ".join(norm(c.text) for c in header_cells[:6])
            except Exception:
                snippet = "(empty/unreadable)"
            log(f"  [table check] candidate (context={name}): {n} rows, headers: {snippet}")
            all_candidates.append((n, name, t))

    if not all_candidates:
        return None
    all_candidates.sort(key=lambda x: x[0], reverse=True)
    best_n, best_name, best_t = all_candidates[0]
    # Re-switch into the context the chosen fallback table actually belongs to before returning it -
    # a stale element reference from an earlier context switch would otherwise fail on use.
    for name, switch in _iter_contexts(driver):
        if name == best_name:
            switch()
            break
    log(f"  Using the table with the most rows (context={best_name}, be sure to verify this)")
    return best_t


# Each header cell's clickable sort-toggle appends "Unsorted"/"Ascending"/"Descending" to its own
# text (confirmed live - "First Name Unsorted", "Last Name Unsorted", etc.) - strip it so the
# Excel header row reads "First Name", not "First Name Unsorted".
SORT_SUFFIX_RE = re.compile(r"\s+(Unsorted|Ascending|Descending)\s*$", re.I)


def _clean_header_text(text):
    return SORT_SUFFIX_RE.sub("", text).strip()


def _extract_table(table):
    """Reads only DIRECT child rows/cells (see _direct_rows/_direct_cells) - table.find_elements(By.
    TAG_NAME, "tr"/"td") descend into EVERY nested table too, and this admin console's report viewer
    wraps every single cell's own text in its own one-cell nested <table> (a per-cell text-formatting
    wrapper, confirmed live via outerHTML - not a "grouped sub-table per user" as first assumed).
    That inner one-cell <tr> has no nested table of ITS OWN, so an earlier "skip rows containing a
    nested table" filter let it through as if it were a genuine standalone row - producing "First
    Name" / "Last Name" / ... each on its own row in column A instead of a normal 40+-column grid."""
    rows_out = []
    for row in _direct_rows(table):
        cells = _direct_cells(row)
        if len(cells) < 10:
            # A short row is usually junk, but it is ALSO how the grid renders its per-person group
            # band: a single full-width cell holding that person's Skillport LOGIN username (e.g.
            # "10lc387010"). That username is the only identifier in this whole report that actually
            # resolves against our own records - the per-row "User ID" column is Skillport's internal
            # numeric account id (e.g. "509876"), which matches nothing on our side. Dropping these
            # rows wholesale (what the bare `continue` used to do) is why every exported file lost
            # the real identifier and forced identity resolution to fall back to name-matching.
            #
            # Emit the band as a one-value row so the exported sheet carries it directly above that
            # person's rows - exactly the group-header shape the importer already carries down.
            band = [norm(c.text) for c in cells]
            band_values = [t for t in band if t]
            if rows_out and len(band_values) == 1 and GROUP_BAND_RE.match(band_values[0]):
                rows_out.append([band_values[0]])
            continue
        text_cells = [norm(c.text) for c in cells]
        if not any(text_cells):
            continue
        if not rows_out:
            rows_out.append([_clean_header_text(t) for t in text_cells])
        else:
            rows_out.append(text_cells)
    return rows_out


def _read_page_label(driver):
    # This admin console is heavily frame-based (see the module's own docstring) - the paging
    # toolbar ("Page [__] of NN") can live in a DIFFERENT frame than whichever one the driver
    # happens to be switched into after finding the results table. Reading only
    # driver.find_element(By.TAG_NAME, "body").text against the CURRENT context silently returns
    # nothing when the toolbar is elsewhere, which is indistinguishable from "only one page" -
    # confirmed live: a real report showing "Page 1 of 33" in the browser was read back as
    # total_pages=None by the old single-context version of this function. Search every frame.
    for name, switch in _iter_contexts(driver):
        try:
            switch()
            body_text = driver.find_element(By.TAG_NAME, "body").text
        except Exception:
            continue

        m_total = re.search(r"of\s+(\d+)\b", body_text, re.I)
        total = int(m_total.group(1)) if m_total else None
        if total is None:
            continue

        # The current-page number sits inside an editable <input> (a "jump to page" box) - its
        # value is never part of body.text, so read the input directly instead of relying on the
        # same regex working against plain text for both numbers.
        page_num = None
        m_plain = PAGE_LABEL_RE.search(body_text)
        if m_plain:
            page_num = int(m_plain.group(1))
            total = int(m_plain.group(2))
        else:
            for inp in driver.find_elements(By.CSS_SELECTOR, "input[type='text'], input:not([type])"):
                val = (inp.get_attribute("value") or "").strip()
                if val.isdigit() and 1 <= int(val) <= total:
                    page_num = int(val)
                    break

        return page_num, total

    driver.switch_to.default_content()
    return None, None


def _click_next_page(driver):
    """Several fallback strategies, since this is an icon-only pagination control I can't verify
    live. Searches EVERY frame (see _read_page_label's comment - the paging toolbar can live in a
    different frame than the results table, and this is called right after _find_results_table
    leaves the driver switched into THAT table's frame, not necessarily the toolbar's). If every
    automated strategy fails and we're not headless, pauses for the user to click it manually."""
    candidates = [
        (By.CSS_SELECTOR, "a[title*='next' i], img[title*='next' i], img[alt*='next' i]"),
        (By.CSS_SELECTOR, "[class*='next' i]"),
        (By.CSS_SELECTOR, "a[onclick*='next' i]"),
    ]
    for name, switch in _iter_contexts(driver):
        try:
            switch()
        except Exception:
            continue
        for by, sel in candidates:
            try:
                els = [e for e in driver.find_elements(by, sel) if e.is_displayed()]
                if els:
                    js_click(driver, els[0])
                    return True
            except Exception:
                continue

        # Positional fallback: the pagination bar is "|<  <  [page-number input]  >  >|" (First,
        # Prev, page box, Next, Last) - the clickable icon right after the page-number input is Next.
        try:
            for inp in driver.find_elements(By.CSS_SELECTOR, "input[type='text'], input:not([type])"):
                if not (inp.get_attribute("value") or "").strip().isdigit():
                    continue
                siblings = driver.execute_script(
                    "let el = arguments[0].nextElementSibling; let out = []; "
                    "while (el && out.length < 4) { out.push(el); el = el.nextElementSibling; } return out;",
                    inp)
                for sib in siblings:
                    if sib.tag_name.lower() in ("a", "button", "img", "span") and sib.is_displayed():
                        js_click(driver, sib)
                        return True
        except Exception:
            pass

    driver.switch_to.default_content()
    return False


class RecoverableScrapeError(RuntimeError):
    """A page that should have rendered and didn't - as opposed to a genuinely broken scrape.

    Confirmed live at page 1922: the paging label already read 1922, so navigation had worked
    perfectly; only the grid never appeared. That is a browser that has degraded (this run was on a
    machine with ~0.5 GB free), not a report that has ended or a selector that is wrong - and the
    right response to it is a fresh browser resuming from the last completed page, not killing a
    run that is 67,000 rows deep.

    Bounded by MAX_BROWSER_RESTARTS, so a genuine, permanent failure still stops rather than looping.
    """


def _is_browser_dead(exc):
    """Is this exception "the browser/ChromeDriver went away", as opposed to a real problem with the
    page or the data?

    Only the first kind is worth relaunching for. Retrying a genuine failure - a missing Preview
    button, an empty table, a bad selector - would just loop forever hitting the same wall, so
    everything not on this list is still fatal.

    Matched on the message rather than the exception type because Selenium wraps urllib3's
    connection errors, and the useful signal ("connection refused", "invalid session id") ends up in
    the text either way.
    """
    text = f"{type(exc).__name__} {exc}".lower()
    markers = (
        "maxretryerror",              # ChromeDriver's HTTP endpoint stopped answering
        "newconnectionerror",
        "connectionrefused",
        "actively refused",           # WinError 10061 - the driver process is gone
        "remotedisconnected",
        "connection aborted",
        "connection reset",
        "readtimeout",                # driver hung long enough to be unusable
        "read timed out",             # same thing worded by urllib3 rather than named by its type
        "invalid session id",         # session died under it
        "no such session",
        "session deleted",
        "chrome not reachable",
        "disconnected: not connected to devtools",
        "target window already closed",
        "web view not found",
    )
    return any(m in text for m in markers)


JUMP_CONFIRM_TIMEOUT = 90


def _jump_commit_strategies():
    """Ways to make the viewer accept a typed page number, tried in order.

    Confirmed live: typing the number and sending RETURN through Selenium put "1819" in the box but
    left the label on page 1 - this is an ExtJS-style field, and ExtJS binds its handler to a raw
    keydown with keyCode 13 rather than to whatever Selenium's send_keys produces. Setting .value
    from JavaScript has the same problem in reverse: the DOM shows the new text but no event ever
    fires, so the widget's internal value never changes.

    Hence several strategies rather than one: each commits the value a different way, and the first
    one the page actually honours wins.
    """

    def native_enter(driver, inp, target):
        inp.click()
        inp.send_keys(Keys.CONTROL, "a")
        inp.send_keys(str(target))
        inp.send_keys(Keys.RETURN)

    def native_tab(driver, inp, target):
        # Many paging widgets commit on blur rather than on Enter.
        inp.click()
        inp.send_keys(Keys.CONTROL, "a")
        inp.send_keys(str(target))
        inp.send_keys(Keys.TAB)

    def js_value_then_enter(driver, inp, target):
        # Set the text through the native value setter so frameworks that patch .value still see it,
        # fire the events a real edit would, then press Enter for real.
        driver.execute_script("""
            const el = arguments[0], value = arguments[1];
            const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(el, value);
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
        """, inp, str(target))
        inp.send_keys(Keys.RETURN)

    def js_value_then_synthetic_keydown(driver, inp, target):
        # ExtJS reads keyCode/which, which modern KeyboardEvent constructors leave at 0 unless they
        # are forced back on afterwards.
        driver.execute_script("""
            const el = arguments[0], value = arguments[1];
            const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(el, value);
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
            for (const type of ['keydown', 'keypress', 'keyup']) {
                const ev = new KeyboardEvent(type, {bubbles: true, cancelable: true, key: 'Enter', code: 'Enter'});
                Object.defineProperty(ev, 'keyCode', {get: () => 13});
                Object.defineProperty(ev, 'which', {get: () => 13});
                el.dispatchEvent(ev);
            }
            el.blur();
        """, inp, str(target))

    return [
        ("native + Enter", native_enter),
        ("JS value + Enter", js_value_then_enter),
        ("JS value + synthetic keydown 13", js_value_then_synthetic_keydown),
        ("native + Tab (commit on blur)", native_tab),
    ]


def _wait_for_page_label(driver, target, timeout):
    """Polls until the paging label reads `target`.

    Deliberately NOT _wait_for_page_table: the previous page's table is still on screen the instant
    a jump is submitted, so waiting for "a table exists" returned true immediately and the jump was
    judged failed before the new page had any chance to load. The label is the only thing that
    actually says which page is showing.
    """
    end = time.time() + timeout
    last_seen = None
    while time.time() < end:
        current, _ = _read_page_label(driver)
        if current == target:
            return True
        if current != last_seen:
            last_seen = current
            log(f"    [jump] label currently {current} - waiting for {target}...")
        time.sleep(1.5)
    return False


def _find_page_box(driver):
    """(context_name, element, total_pages) for the viewer's jump-to-page input, or (None, None, None).

    Identified by holding a number that is a plausible current page, which is how _read_page_label
    already finds the same box.
    """
    for name, switch in _iter_contexts(driver):
        try:
            switch()
            body_text = driver.find_element(By.TAG_NAME, "body").text
        except Exception:
            continue

        m_total = re.search(r"of\s+(\d+)\b", body_text, re.I)
        if not m_total:
            continue
        total = int(m_total.group(1))

        for inp in driver.find_elements(By.CSS_SELECTOR, "input[type='text'], input:not([type]), input[type='number']"):
            try:
                val = (inp.get_attribute("value") or "").strip()
                if val.isdigit() and 1 <= int(val) <= total and inp.is_displayed() and inp.is_enabled():
                    return name, inp, total
            except Exception:
                continue

    driver.switch_to.default_content()
    return None, None, None


def _jump_to_page(driver, target):
    """Types the page number straight into the viewer's "Page [__] of N" box.

    Turns a 1,800-page resume from ~2.5 hours of clicking Next into seconds. Returns True only once
    the label CONFIRMS the target page - an unverified jump would land somewhere else entirely and
    scrape the wrong part of the report while looking perfectly healthy.
    """
    context_name, inp, total = _find_page_box(driver)
    if inp is None:
        log("  [jump] page number box was not found")
        return False

    if target > total:
        log(f"  [jump] target {target} is greater than the total pages {total}")
        return False

    log(f"  [jump] found the page box (context={context_name}) - jumping to {target}")

    for attempt, (label, commit) in enumerate(_jump_commit_strategies()):
        try:
            # Re-find between attempts: a partly-successful attempt can re-render the toolbar and
            # leave the previous element handle stale. (Comparing function identity here would not
            # work - _jump_commit_strategies() builds fresh closures on every call.)
            if attempt > 0:
                context_name, inp, total = _find_page_box(driver)
                if inp is None:
                    log("  [jump] could not find the box again")
                    return False

            log(f"    [jump] trying: {label}")
            commit(driver, inp, target)

            if _wait_for_page_label(driver, target, JUMP_CONFIRM_TIMEOUT):
                # The label is right; make sure the grid itself has actually rendered before the
                # caller starts scraping it.
                if _wait_for_page_table(driver, target):
                    log(f"  [jump] reached page {target} ({label})")
                    return True
                log("  [jump] the label is right but the table did not appear")
                return False

            log(f"    [jump] {label} did not change the page")
        except Exception as e:
            log(f"    [jump] problem with {label}: {e}")

    _dump_page_box_debug(driver)
    return False


def _dump_page_box_debug(driver):
    """Logs every candidate input's identifying attributes. If all four commit strategies failed,
    this is what says which widget the toolbar actually uses - without another blind guess."""
    log("  [jump] none of the strategies worked - here are the toolbar's inputs:")
    for name, switch in _iter_contexts(driver):
        try:
            switch()
            inputs = driver.find_elements(By.TAG_NAME, "input")
        except Exception:
            continue
        for inp in inputs[:12]:
            try:
                if not inp.is_displayed():
                    continue
                log(f"    [jump debug] context={name} "
                    f"type={inp.get_attribute('type')!r} id={inp.get_attribute('id')!r} "
                    f"name={inp.get_attribute('name')!r} class={inp.get_attribute('class')!r} "
                    f"value={inp.get_attribute('value')!r}")
            except Exception:
                continue
    driver.switch_to.default_content()


def _skip_forward_to_page(driver, start_page, limit):
    """Clicks Next until the grid is showing start_page, WITHOUT scraping anything on the way.

    This is the price of resuming: the report viewer has no "jump to page N", so the only way back
    to where a dead run stopped is to page through. Still far cheaper than re-scraping - skipping a
    page is one click and a short wait, scraping one is a full table extraction.
    """
    # Start from where the viewer ACTUALLY is, not from an assumed page 1. A restored report
    # session can come back on any page, and blindly clicking Next (start_page - 1) times from
    # there would overshoot and silently skip real data.
    current, _ = _read_page_label(driver)
    if current is None:
        log("Resume: could not read the page label - assuming forward from page 1")
        current = 1

    if current == start_page:
        log(f"Resume: the viewer is already on page {start_page} - no need to move forward")
        return

    if current > start_page:
        raise RuntimeError(
            f"Resume: the viewer is on page {current}, which is past the target page {start_page}. The "
            f"report viewer can only move forward - either start over (without --resume) or navigate the "
            f"browser to page {start_page} yourself and pass --start-page.")

    # Try the jump-to-page box first - seconds instead of hours. Falls back to clicking Next when
    # the box is missing or the jump can't be confirmed, so this is a pure speed-up: it never
    # changes WHERE we end up, only how long getting there takes.
    log(f"Resume: trying to jump to page {start_page}...")
    if _jump_to_page(driver, start_page):
        log(f"Resume: reached page {start_page} (via jump) - starting the scrape now")
        return

    log("Resume: the jump did not work - moving forward by clicking 'Next' (this will take longer)")
    log(f"Resume: moving forward from page {current} to {start_page} (without scraping)...")
    skip_started = time.time()
    total_to_skip = start_page - current

    for done, p in enumerate(range(current, start_page), start=1):
        if not _click_next_page(driver):
            raise RuntimeError(
                f"Resume: 'Next' was not found on page {p} - could not reach page {start_page}.")

        if not _wait_for_page_table(driver, p + 1):
            shot(driver, f"resume_stuck_page_{p + 1}")
            raise RuntimeError(f"Resume: page {p + 1} did not load.")

        if (done % 50) == 0:
            elapsed = time.time() - skip_started
            remaining = int((elapsed / done) * (total_to_skip - done))
            log(f"  [resume] skipped {done}/{total_to_skip} pages, "
                f"{int(elapsed) // 60}m elapsed, ~{remaining // 60}m remaining")

    current, _ = _read_page_label(driver)
    if current is not None and current != start_page:
        raise RuntimeError(
            f"Resume: after moving forward the label shows page {current}, expected {start_page}. "
            f"Stopping so the wrong pages don't get scraped.")

    log(f"Resume: reached page {start_page}, starting the scrape now")


def save_progress(last_completed_page, row_count):
    """Written after every page so a killed run knows exactly where to resume from. Row COUNT is
    only for the log - the rows themselves are already in the checkpoint CSV."""
    try:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        OUTPUT_PROGRESS_JSON.write_text(
            json.dumps({"last_completed_page": last_completed_page, "rows": row_count}),
            encoding="utf-8")
    except Exception as e:
        log(f"  WARNING: problem writing the progress file: {e}")


def load_progress():
    """(last_completed_page, rows) from a previous run, or (0, []) when there is nothing to resume."""
    if not OUTPUT_PROGRESS_JSON.exists():
        return 0, None, None

    try:
        state = json.loads(OUTPUT_PROGRESS_JSON.read_text(encoding="utf-8"))
        last_page = int(state.get("last_completed_page", 0))
    except Exception as e:
        log(f"WARNING: could not read the progress file ({e}) - starting from the beginning")
        return 0, None, None

    if last_page < 1 or not OUTPUT_CSV_CHECKPOINT.exists():
        return 0, None, None

    import csv
    with open(OUTPUT_CSV_CHECKPOINT, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    if not rows:
        return 0, None, None

    return last_page, rows[0], rows[1:]


def scrape_all_pages(driver, max_pages=None, start_page=1, resume_rows=None, resume_headers=None):
    """start_page > 1 pages FORWARD without scraping, then carries on from there - the resume path
    after a run dies mid-file. resume_rows/resume_headers carry the already-scraped rows back in so
    the final export is the whole report, not just the tail."""
    all_rows = list(resume_rows or [])
    headers = resume_headers
    page_num, total_pages = _read_page_label(driver)
    if total_pages is None:
        log("WARNING: the 'Page X of Y' label was not found - only the current page will be scraped")
        total_pages = 1
    log(f"Total pages: {total_pages}")

    limit = min(total_pages, max_pages) if max_pages else total_pages

    if start_page > 1:
        _skip_forward_to_page(driver, start_page, limit)

    run_started = time.time()

    for p in range(start_page, limit + 1):
        # An All Groups / 1-year run is thousands of pages and takes hours. Report elapsed time and a
        # projected finish every 25 pages so a long run is visibly progressing and can be planned
        # around, rather than looking like it has stalled. The per-page checkpoint below means
        # stopping it early never loses what has already been scraped.
        if p % 25 == 0:
            elapsed = time.time() - run_started
            done_this_run = max(p - start_page, 1)
            remaining = int((elapsed / done_this_run) * (limit - p))
            log(f"  [progress] {p}/{limit} pages, {len(all_rows)} rows, "
                f"{int(elapsed) // 60}m elapsed, ~{remaining // 60}m remaining")

        log(f"Scraping page {p}/{limit}...")
        # strict from page 2 on: only page 1 may fall back to the largest candidate table (that
        # fallback exists for the first render, where the frame layout is still being learned).
        # Mid-run it would silently substitute toolbar junk for a not-yet-rendered grid.
        table = _find_results_table(driver, strict=p > 1)
        if table is None:
            shot(driver, f"no_table_page_{p}")
            raise RecoverableScrapeError(
                f"No results table was found on page {p}/{limit} ({len(all_rows)} rows preserved).")

        extracted = _extract_table(table)
        if not extracted:
            # Only the LAST page may legitimately come back empty. Anywhere else this means the
            # grid hadn't finished rendering - the old code broke out here and saved a partial
            # export as if the report had simply ended.
            if p < limit:
                shot(driver, f"empty_table_page_{p}")
                # Recoverable, not fatal: mid-report this means the grid rendered empty, which a
                # fresh browser normally fixes. Accepting it as "the end" would silently truncate
                # the export - the failure this whole guard exists to prevent.
                raise RecoverableScrapeError(
                    f"The table on page {p}/{limit} came back empty while {limit - p} pages were still "
                    f"remaining. {len(all_rows)} rows preserved.")
            log(f"  Page {p}: table came back empty (last page) - stopping here")
            break

        if headers is None:
            headers = extracted[0]
            all_rows.extend(extracted[1:])
        else:
            # Skip a repeated header row if this page's table starts with one again.
            body_rows = extracted[1:] if extracted[0] == headers else extracted
            all_rows.extend(body_rows)

        log(f"  Page {p}: found {len(extracted) - 1} rows (total so far: {len(all_rows)})")

        # Checkpoint after every page - if the script crashes or is stopped partway through, the
        # data scraped so far will not be lost.
        save_checkpoint_csv(headers or [], all_rows)
        save_progress(p, len(all_rows))

        if p == limit:
            break

        if not _click_next_page(driver):
            log("  Could not find the 'Next' button automatically.")
            try:
                input("  Click the 'Next' page arrow manually in the browser, then press Enter here "
                      "(or just press Enter if this is the last page)...")
            except EOFError:
                log("  Manual intervention is not possible in headless mode - stopping here.")
                break

        # Same reasoning as the post-Preview wait - don't scrape the next page until a real,
        # header-validated table actually exists again (the grid briefly disappears/reloads after
        # clicking Next).
        #
        # This used to wait 30s and then carry on regardless ("trying anyway"), which
        # is how a 1302-page run ended at page 21 with 690 rows and exit code 0: the grid was still
        # loading, the non-strict table search grabbed a blank 3-row toolbar table, that extracted
        # as empty, and the loop treated empty as end-of-data. A slow page is now waited out, and a
        # page that genuinely never arrives is an error, not a short export.
        if not _wait_for_page_table(driver, p + 1):
            shot(driver, f"page_{p + 1}_never_loaded")
            raise RecoverableScrapeError(
                f"The results table for page {p + 1}/{limit} did not load within {PAGE_LOAD_TIMEOUT}s "
                f"(even after 3 attempts). {len(all_rows)} rows preserved.")

        new_page_num, _ = _read_page_label(driver)
        if new_page_num is not None and new_page_num == p:
            log("  The page number did not change - the 'Next' click probably did not work, stopping")
            break

    return headers or [], all_rows


def _wait_for_page_table(driver, page_number):
    """Waits for a header-VALIDATED results table to appear after a Next click, up to
    PAGE_LOAD_TIMEOUT per attempt, 3 attempts. Returns True once the table is there.

    On a lapsed wait it re-clicks Next ONLY when the paging label proves the previous click never
    registered. Re-clicking unconditionally would silently skip a whole page whenever the click had
    in fact registered and the grid was merely slow - trading a visible stall for invisible missing
    rows. If the label can't be read, it waits again rather than risking the skip."""
    for attempt in range(3):
        end = time.time() + PAGE_LOAD_TIMEOUT
        last_beat = time.time()
        while time.time() < end:
            if any_context_has_valid_table(driver):
                return True
            if time.time() - last_beat >= 30:
                last_beat = time.time()
                log(f"    ...page {page_number} is still loading "
                    f"({int(end - time.time())}s remaining)")
            time.sleep(1.0)

        if attempt == 2:
            break

        current, _ = _read_page_label(driver)
        if current is not None and current < page_number:
            log(f"  WARNING: the table for page {page_number} did not appear within {PAGE_LOAD_TIMEOUT}s and "
                f"the label is still showing page {current} - clicking 'Next' again "
                f"(attempt {attempt + 2}/3)")
            _click_next_page(driver)
        else:
            log(f"  WARNING: the table for page {page_number} did not appear within {PAGE_LOAD_TIMEOUT}s "
                f"(label: {current}) - 'Next' was already clicked, just waiting longer "
                f"(attempt {attempt + 2}/3)")

    return False


def save_checkpoint_csv(headers, rows):
    import csv
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_CSV_CHECKPOINT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        if headers:
            w.writerow(headers)
        w.writerows(rows)


def save_to_excel(headers, rows, path):
    """Writes the workbook in openpyxl's WRITE-ONLY (streaming) mode.

    The ordinary Workbook builds every cell as a Python object in memory before saving anything. A
    full All-Groups run is ~105,000 rows x 43 columns - about 4.5 MILLION cells, which costs well
    over a gigabyte. This runs on a machine that has been sitting at under 1 GB free all night, so
    the ordinary path would raise MemoryError at the very END of a seven-hour scrape, with every row
    already collected and nothing written out.

    Write-only mode streams each row straight to the file instead, so peak memory stays flat no
    matter how many rows there are. The header still gets its styling, and freeze panes/filter/column
    widths are set explicitly because ws.dimensions is not known ahead of time in this mode.
    """
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    column_count = len(headers) or max((len(r) for r in rows), default=1)

    def build():
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Asset Activity by User")

        for i in range(1, column_count + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        if headers:
            fill = PatternFill("solid", start_color="1F2937")
            header_cells = []
            for value in headers:
                cell = WriteOnlyCell(ws, value=value)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
                header_cells.append(cell)
            ws.append(header_cells)

            ws.freeze_panes = "A2"
            # Computed, not taken from ws.dimensions - a write-only sheet does not know its extent
            # until it has been saved.
            ws.auto_filter.ref = f"A1:{get_column_letter(column_count)}{len(rows) + 1}"

        for r in rows:
            ws.append(list(r))

        return wb

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    try:
        build().save(path)
        log(f"Excel file created: {path} ({len(rows)} rows)")
    except PermissionError:
        # File probably open in Excel right now (Windows locks it for other processes) - don't
        # lose all the scraped data over this, save under a timestamped name instead. Rebuilt
        # because a write-only workbook cannot be saved twice.
        alt_path = Path(path).with_stem(f"{Path(path).stem}_{time.strftime('%Y%m%d_%H%M%S')}")
        log(f"WARNING: '{path}' is locked (probably open in Excel) - saving to '{alt_path}' instead")
        build().save(alt_path)
        log(f"Excel file created: {alt_path} ({len(rows)} rows)")


def save_to_notepad(headers, rows, path):
    """Plain, human-readable .txt export - same data as the Excel file, laid out one record per
    block instead of a grid, matching the existing skillport_scraper.py's export_notepad() style.

    Streamed to disk rather than assembled in memory: a full run is ~105,000 records of ~45 lines
    each, and holding those ~4.7 million strings in a list before joining them is a second way to
    run out of memory at the end of a long scrape.
    """
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    separator = "=" * 90

    with open(path, "w", encoding="utf-8") as f:
        for i, row in enumerate(rows, 1):
            f.write(f"{separator}\nRecord {i}\n{separator}\n")
            for col_name, value in zip(headers, row):
                f.write(f"{col_name:<30}: {value}\n")
            f.write("\n")

    log(f"Notepad file created: {path} ({len(rows)} rows)")


# ===================== ENTRY =====================

def main():
    # Declared up here, not next to the assignment below: Python requires the `global` statement to
    # precede EVERY use of the name in the function, and the argument default a few lines down reads
    # it (SyntaxError otherwise).
    global REPORT_LOAD_TIMEOUT, PAGE_LOAD_TIMEOUT

    ap = argparse.ArgumentParser()
    ap.add_argument("--headless", action="store_true",
                     help="Don't pass this the first time - keep the browser open to verify the selectors.")
    ap.add_argument("--max-pages", type=int, default=None, help="Only N pages, for testing.")
    ap.add_argument("--group", default=None,
                     help="One-off Group Name override - the default comes from the Settings table "
                          "(or DEFAULT_GROUP_NAME if that isn't found either).")
    ap.add_argument("--report-load-timeout", type=int, default=REPORT_LOAD_TIMEOUT,
                     help="Max wait (seconds) for the report to render after Preview. Default "
                          f"{REPORT_LOAD_TIMEOUT}. Increase for a large report like All Groups + 1 year.")
    ap.add_argument("--resume", action="store_true",
                     help="Continue a previous unfinished run from where it left off (from the progress "
                          "file + checkpoint CSV). If a 3000-page run dies from the internet going down, "
                          "this preserves all of it.")
    ap.add_argument("--start-page", type=int, default=None,
                     help="Start scraping from this page (earlier pages are only skipped, not "
                          "scraped). --resume sets this itself; only pass it manually when there is no "
                          "progress file.")
    ap.add_argument("--page-load-timeout", type=int, default=PAGE_LOAD_TIMEOUT,
                     help="Max wait (seconds) for each subsequent page's grid to appear after 'Next'. Default "
                          f"{PAGE_LOAD_TIMEOUT}, up to 3 attempts. Increase if the server is slow.")
    ap.add_argument("--date-amount", type=int, default=1,
                     help="Activity Dates 'Previous N' - default 1.")
    ap.add_argument("--date-unit", default="Year",
                     help="Activity Dates unit: Year / Month / Week / Day. Default Year, because the "
                          "template's own default of one month silently drops older history from the export.")
    ap.add_argument("--date-from", default=None,
                     help="Activity Dates start of an explicit range, as dd/mm/yyyy. Passing this "
                          "(together with --date-to) switches to the 'Custom' range mode and "
                          "overrides --date-amount/--date-unit entirely.")
    ap.add_argument("--date-to", default=None,
                     help="Activity Dates end of an explicit range, as dd/mm/yyyy. See --date-from.")
    ap.add_argument("--output", default=None,
                     help="Path to the output .xlsx. Required for multi-group runs - otherwise every "
                          "group overwrites the same file and only the last group survives. "
                          "The .txt companion file is derived from this same name.")
    args = ap.parse_args()

    REPORT_LOAD_TIMEOUT = args.report_load_timeout
    PAGE_LOAD_TIMEOUT = args.page_load_timeout

    if PASSWORD == "PUT_PASSWORD_HERE":
        log("Set the SKILLPORT_PASS environment variable, or edit PASSWORD in skillport_scraper.py.")
        sys.exit(1)

    group_name = args.group or _fetch_group_name_from_db()

    # Resume state is worked out BEFORE the browser opens, so a bad/missing progress file fails
    # fast instead of after twenty minutes of logging in and re-running the report.
    start_page, resume_headers, resume_rows = 1, None, None
    if args.resume:
        last_page, resume_headers, resume_rows = load_progress()
        if last_page < 1:
            log("--resume was given but no previous progress was found - starting from the beginning")
        else:
            start_page = last_page + 1
            log(f"RESUME: the previous run got to page {last_page} ({len(resume_rows or [])} rows "
                f"in the checkpoint) - continuing from page {start_page}")
    if args.start_page:
        start_page = args.start_page
        log(f"--start-page {start_page} was given")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    headers = rows = None
    attempt = 0

    # Supervisor loop. A 3000-page report keeps one Chrome tab alive for hours while an ExtJS grid
    # accumulates DOM, and on a machine under memory pressure the browser eventually gets killed -
    # which used to end the whole run with a ConnectionRefused stack trace after 1700 pages of work.
    # Progress is written after EVERY page, so the honest response to a dead browser is to start a
    # fresh one and carry on from the last saved page, not to give up.
    while True:
        driver = None
        try:
            driver = make_driver(args.headless)
            login(driver)

            # On a resume, Skillport often restores the previous report session and drops us
            # straight into the VIEWER. That report was generated with the settings we want anyway,
            # so re-opening the template editor and re-running it would throw away the very thing we
            # are trying to continue - and cost another hour regenerating 3000 pages.
            if start_page > 1 and any_context_has_valid_table(driver):
                log("Resume: the previous report is still open - not reconfiguring the template")
            else:
                open_asset_activity_template(driver)
                configure_group_filter(driver, group_name)
                if args.date_from and args.date_to:
                    configure_activity_dates_range(driver, args.date_from, args.date_to)
                else:
                    configure_activity_dates(driver, args.date_amount, args.date_unit)
                expand_and_select_all_display_options(driver)

            click_preview(driver, allow_existing_report=start_page > 1)
            headers, rows = scrape_all_pages(
                driver, args.max_pages, start_page=start_page,
                resume_rows=resume_rows, resume_headers=resume_headers)
            break

        except KeyboardInterrupt:
            log("Stopped - the checkpoint CSV (output/asset_activity_by_user_checkpoint.csv) is preserved. "
                "You can continue from there with `--resume`.")
            return

        except Exception as e:
            # Two different failures deserve the same response - a fresh browser resuming where the
            # last one stopped. Either the driver went away outright, or it stayed up but stopped
            # rendering (page 1922 live: the label had advanced correctly, the grid never appeared).
            crashed = _is_browser_dead(e)
            stalled = isinstance(e, RecoverableScrapeError)

            if not crashed and not stalled:
                log(f"FATAL: {e}")
                shot(driver, "fatal")
                raise

            attempt += 1
            if attempt > MAX_BROWSER_RESTARTS:
                log(f"FATAL: still not working after {MAX_BROWSER_RESTARTS} restarts - stopping. "
                    f"The checkpoint is preserved, continue with `--resume`. Last error: {e}")
                raise

            # Re-read progress from DISK, not from memory: the in-memory counters belong to the
            # attempt that just died, while the progress file is the last page actually persisted.
            last_page, resume_headers, resume_rows = load_progress()
            start_page = max(last_page + 1, start_page)
            log(f"{'BROWSER DIED' if crashed else 'PAGE DID NOT RENDER'} ({type(e).__name__}): {e}")
            log(f"  -> opening a new browser and continuing from page {start_page} "
                f"(attempt {attempt}/{MAX_BROWSER_RESTARTS}, {len(resume_rows or [])} rows preserved).")

        finally:
            if driver is not None:
                try:
                    driver.quit()
                except Exception:
                    force_kill_driver(driver)

    if not rows:
        log("No data was scraped.")
        return

    out_xlsx = Path(args.output) if args.output else OUTPUT_XLSX
    out_txt = out_xlsx.with_suffix(".txt")
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    # Each export is attempted independently. These run at the very END of a scrape that can take
    # seven hours, so one of them failing must not take the other down with it - and either way the
    # checkpoint CSV already holds every row, which is the thing worth saying out loud rather than
    # leaving the operator to work out from a traceback.
    wrote_any = False
    for label, writer, target in [
        ("Excel", save_to_excel, out_xlsx),
        ("Notepad", save_to_notepad, out_txt),
    ]:
        try:
            writer(headers, rows, target)
            wrote_any = True
        except Exception as e:
            log(f"WARNING: the {label} file could not be created ({type(e).__name__}: {e}). "
                f"All the data is still preserved in {OUTPUT_CSV_CHECKPOINT} ({len(rows)} rows).")

    if not wrote_any:
        log(f"Both exports failed - but all {len(rows)} rows are fully preserved in the checkpoint CSV: "
            f"{OUTPUT_CSV_CHECKPOINT}")


if __name__ == "__main__":
    main()
