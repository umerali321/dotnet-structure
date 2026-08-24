#!/usr/bin/env python3
"""
SkillSets Online (Skillport) course scraper bot  v8
Click-driven scraping for the ExtJS-based site.

Modes:
  --mode top      only the category page's own Courses list
  --mode sidebar  every sub-category in the left menu
  --mode both     both

Category:
  --category "PMI COURSES"    a single category
  --category ALL              every top category in one run

Install:  pip install selenium openpyxl
Run:
  set SKILLPORT_PASS=SkillTalha$26
  python skillport_scraper.py --category ALL --mode sidebar
  python skillport_scraper.py --category ALL --mode both
  python skillport_scraper.py --category ALL --mode sidebar --limit 2
  python skillport_scraper.py --category ALL --mode sidebar --no-section-urls --no-launch
  python skillport_scraper.py --category ALL --mode sidebar --export-only
  python skillport_scraper.py --debug
"""

import argparse, json, os, re, subprocess, time, traceback
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import (
    ElementClickInterceptedException, StaleElementReferenceException, WebDriverException)

# ===================== CONFIG =====================
BASE      = "https://skillsetsonline.skillport.com/skillportfe"
LOGIN_URL = f"{BASE}/login.action"
MAIN_URL  = f"{BASE}/main.action"

USERNAME = os.environ.get("SKILLPORT_USER", "talha_admin")
PASSWORD = os.environ.get("SKILLPORT_PASS", "SkillTalha$26")

DEFAULT_CATEGORY = "MANAGEMENT COURSES"

ALL_CATEGORIES = [
    "MANAGEMENT COURSES",
    "COMMUNICATION COURSES",
    "AI IN THE WORKPLACE COURSES",
    "CUSTOMER SERVICE, SALES & MARKETING COURSES",
    "PRIVACY & CYBERSECURITY COURSES",
    "MICROSOFT / DESKTOP COURSES",
    "PMI COURSES",
    "SIX SIGMA COURSES",
    "COMPLIANCE COURSES",
    "HR PROFESSIONAL COURSES",
    "BUDGET PLANNING COURSES",
    "OTHER COURSES",
    "COURSES FOR IT PROFESSIONALS",
    "LABS FOR IT PROFESSIONALS",
]

# TypeID 1 = NON_IT, TypeID 2 = IT (matches GetCourseLibraryQueryHandler.MapType on the backend).
IT_CATEGORY_NAMES = {"COURSES FOR IT PROFESSIONALS", "LABS FOR IT PROFESSIONALS"}

ALL_CATEGORIES_UPPER = {c.strip().upper() for c in ALL_CATEGORIES}

def type_id_for_category(category_name):
    return 2 if (category_name or "").strip().upper() in IT_CATEGORY_NAMES else 1

OUT_DIR      = Path("output")
WAIT         = 25
PAGE_SETTLE  = 3.0
MAX_VIEW_MORE = 60
SIDEBAR_MAX_X = 580

ASSET_SEL          = ".sp-assetContainer.sp-panel-long"
ASSET_SEL_FALLBACK = ".sp-assetContainer:not(.sp-panel-small)"

SKIP_SIDEBAR = []
# Every one of these is a piece of persistent site chrome (top nav / global widgets) with fixed,
# short text - not something that can legitimately appear as a substring of a real category name.
# All exact-match only, same fix as "courses" originally got on its own: a substring check here
# would silently drop real categories that merely happen to contain one of these words, e.g. a
# "Search Engine Marketing" or "Home Office Productivity" or "Business English" sub-category
# would vanish from the captured sidebar list entirely, confused with the "Search"/"Home"/
# "English" nav chrome elsewhere on the page.
SKIP_SIDEBAR_EXACT = ["the library", "my profile", "log out", "learning transcript",
                       "manage languages", "english", "accessibility", "view more",
                       "search", "home", "skillsets", "sub-categories",
                       "categories", "courses"]
# ==================================================


def log(m): print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)
def norm(t): return re.sub(r"\s+", " ", t).strip() if t else ""

def file_tag(mode, category=None):
    base = {"top": "main", "sidebar": "subcategories", "both": "all"}.get(mode, mode)
    if category and category.strip().upper() != "ALL":
        slug = re.sub(r"[^a-z0-9]+", "_", category.lower()).strip("_")[:40]
        return f"{base}_{slug}"
    return base

def clean_lines(t):
    if not t: return ""
    lines = [re.sub(r"[ \t]+", " ", l).strip() for l in t.splitlines()]
    return "\n".join([l for l in lines if l])

def load_json(p, d):
    p = Path(p)
    if p.exists():
        try: return json.loads(p.read_text(encoding="utf-8"))
        except Exception: log(f"WARN: {p} failed to parse")
    return d

def save_json(p, data):
    Path(p).parent.mkdir(parents=True, exist_ok=True)
    Path(p).write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

def shot(driver, name):
    try:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        driver.switch_to.default_content()
        driver.save_screenshot(str(OUT_DIR / f"debug_{name}_{int(time.time())}.png"))
    except Exception: pass


def make_driver(headless=False):
    o = Options()
    if headless: o.add_argument("--headless=new")
    for a in ["--window-size=1600,1000", "--disable-gpu", "--no-sandbox",
              "--disable-dev-shm-usage", "--disable-blink-features=AutomationControlled"]:
        o.add_argument(a)
    o.add_experimental_option("excludeSwitches", ["enable-automation"])
    d = webdriver.Chrome(options=o)
    d.set_page_load_timeout(120)
    return d


def force_kill_driver(driver):
    """driver.quit() asks the browser to close over the same WebDriver connection that just
    died (invalid session, dropped pipe, etc.) - on an already-broken session this silently does
    nothing, leaving the real chromedriver.exe/chrome.exe OS processes running forever as
    zombies. Over a long run with several crashes, enough accumulated zombies can make even a
    freshly-launched replacement browser unstable. Force-kill the actual process tree as a
    backstop so recovery is reliable regardless of why the session died."""
    pid = None
    try:
        pid = driver.service.process.pid
    except Exception:
        pass
    try:
        driver.quit()
    except Exception:
        pass
    if pid:
        try:
            subprocess.run(["taskkill", "/F", "/T", "/PID", str(pid)],
                           capture_output=True, timeout=10)
        except Exception:
            pass


def js_click(driver, el):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    time.sleep(0.3)
    try: el.click()
    except (ElementClickInterceptedException, WebDriverException):
        driver.execute_script("arguments[0].click();", el)


def find_any(driver, locators, timeout=8):
    end = time.time() + timeout
    while time.time() < end:
        for by, sel in locators:
            try:
                for e in driver.find_elements(by, sel):
                    if e.is_displayed(): return e
            except Exception: continue
        time.sleep(0.4)
    return None


def by_text(tag, text):
    t = text.lower()
    low = ("translate(normalize-space(.),"
           "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')")
    return (By.XPATH,
            f"//{tag}[contains({low},\"{t}\") and not(.//{tag}[contains({low},\"{t}\")])]")


# ===================== LOGIN =====================

def page_ready(driver, timeout=30):
    end = time.time() + timeout
    while time.time() < end:
        try:
            if driver.execute_script("return document.readyState") == "complete": return True
        except Exception: pass
        time.sleep(0.5)
    return False


def is_logged_in(driver):
    try: return "login.action" not in driver.current_url.lower()
    except Exception: return False


def switch_to_login_frame(driver):
    driver.switch_to.default_content()
    if driver.find_elements(By.CSS_SELECTOR, "input[type='password']"): return True
    for fr in driver.find_elements(By.CSS_SELECTOR, "iframe, frame"):
        try:
            driver.switch_to.frame(fr)
            if driver.find_elements(By.CSS_SELECTOR, "input[type='password']"): return True
            driver.switch_to.default_content()
        except Exception:
            driver.switch_to.default_content()
    return False


def safe_type(driver, el, text):
    try:
        el.clear(); el.send_keys(text)
        if (el.get_attribute("value") or "") == text: return True
    except StaleElementReferenceException: return False
    except Exception: pass
    try:
        driver.execute_script(
            "arguments[0].value=arguments[1];"
            "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
            "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", el, text)
        return True
    except Exception: return False


def find_login_fields(driver):
    u = find_any(driver, [(By.NAME, "userName"), (By.NAME, "username"),
                          (By.ID, "userName"), (By.ID, "username"),
                          (By.CSS_SELECTOR, "input[type='text']"),
                          (By.CSS_SELECTOR, "input[type='email']")], 12)
    p = find_any(driver, [(By.NAME, "password"), (By.ID, "password"),
                          (By.CSS_SELECTOR, "input[type='password']")], 12)
    return u, p


def login(driver, attempts=3):
    for attempt in range(1, attempts + 1):
        log(f"Login attempt {attempt}/{attempts}")
        try:
            driver.get(LOGIN_URL); page_ready(driver); time.sleep(4.0)
            if is_logged_in(driver):
                log("Already logged in"); return

            switch_to_login_frame(driver)
            u, p = find_login_fields(driver)
            if not u or not p:
                shot(driver, f"login_fields_{attempt}"); time.sleep(3); continue

            if not safe_type(driver, u, USERNAME):
                u, p = find_login_fields(driver)
                if not u or not safe_type(driver, u, USERNAME):
                    time.sleep(3); continue
            _, p = find_login_fields(driver)
            if not p or not safe_type(driver, p, PASSWORD):
                time.sleep(3); continue

            time.sleep(0.8)
            submitted = False
            btn = find_any(driver, [(By.CSS_SELECTOR, "button[type='submit']"),
                                    (By.CSS_SELECTOR, "input[type='submit']"),
                                    by_text("button", "log in"), by_text("button", "login")], 5)
            if btn:
                try: js_click(driver, btn); submitted = True
                except Exception: pass
            if not submitted:
                try:
                    _, p2 = find_login_fields(driver)
                    if p2: p2.send_keys(Keys.RETURN); submitted = True
                except Exception: pass
            if not submitted:
                try:
                    driver.execute_script(
                        "var f=arguments[0].form||document.forms[0]; if(f) f.submit();", p)
                except Exception: pass

            end = time.time() + WAIT
            while time.time() < end and not is_logged_in(driver): time.sleep(1)
            driver.switch_to.default_content(); time.sleep(PAGE_SETTLE)
            if is_logged_in(driver):
                log("Login successful"); return
            shot(driver, f"login_failed_{attempt}")
        except StaleElementReferenceException:
            log("  page reloaded, retrying")
        except Exception as e:
            log(f"  attempt error: {e}"); shot(driver, f"login_error_{attempt}")
        time.sleep(3)
    raise RuntimeError("Login failed 3 times.")


# ===================== NAVIGATION =====================

def goto(driver, url, settle=PAGE_SETTLE, hard=False):
    driver.switch_to.default_content()
    cur = driver.current_url
    if hard or "#" not in url or cur.split("#")[0] != url.split("#")[0]:
        driver.get(url)
    else:
        driver.execute_script("window.location.hash = arguments[0];", url.split("#", 1)[1])
    page_ready(driver)
    time.sleep(settle)


def hard_reload(driver, url, settle=PAGE_SETTLE):
    """Plain, direct re-navigation to url. A previous version of this function tried to detour
    through MAIN_URL first when it thought the browser was already sitting on url - but that
    check compared only the part before '#', which is identical for every page on this site
    (only the hash differs), so the detour fired on literally every call and made sidebar clicks
    fail 100% of the time instead of helping. Reverted - this is just driver.get(url) plus the
    usual settle wait, same as the original goto(..., hard=True) call this replaced."""
    driver.switch_to.default_content()
    driver.get(url)
    page_ready(driver)
    time.sleep(settle)


def open_library(driver):
    log("Opening The Library")
    goto(driver, MAIN_URL, hard=True)
    el = find_any(driver, [by_text("a", "the library"), by_text("span", "the library"),
                           by_text("button", "the library"), by_text("div", "the library")], WAIT)
    if not el:
        shot(driver, "no_library"); raise RuntimeError("The Library link not found")
    js_click(driver, el); time.sleep(PAGE_SETTLE)


def click_top_category(driver, name):
    log(f"Top category: {name}")
    el = find_any(driver, [by_text("a", name.lower()), by_text("span", name.lower()),
                           by_text("li", name.lower()), by_text("div", name.lower())], WAIT)
    if not el:
        shot(driver, "no_category"); raise RuntimeError(f"'{name}' not found")
    js_click(driver, el); time.sleep(PAGE_SETTLE + 2)


JS_SIDEBAR = r"""
const skip = arguments[0], exactSkip = arguments[1], maxx = arguments[2];
const out = [];
document.querySelectorAll('a').forEach(a=>{
  const t = (a.innerText||'').trim();
  if(!t || t.length < 3) return;
  const r = a.getBoundingClientRect();
  if(r.width === 0 && r.height === 0) return;
  if(r.left > maxx) return;
  if(r.top + window.scrollY < 220) return;
  const low = t.toLowerCase();
  if(exactSkip.includes(low)) return;
  if(skip.some(s => low.includes(s))) return;
  out.push(t);
});
return [...new Set(out)];
"""

def get_sidebar_categories(driver):
    driver.execute_script("window.scrollTo(0,0);"); time.sleep(1.0)
    try:
        cats = driver.execute_script(JS_SIDEBAR, SKIP_SIDEBAR, SKIP_SIDEBAR_EXACT, SIDEBAR_MAX_X) or []
    except Exception as e:
        log(f"sidebar read error: {e}"); cats = []
    cats = [norm(c) for c in cats if norm(c)]
    # Some category pages render a "browse other libraries" cross-link widget in the same
    # left-hand region as the real sub-category list - seen live under COURSES FOR IT
    # PROFESSIONALS, whose sidebar read included "PMI COURSES", "SIX SIGMA COURSES", "COMPLIANCE
    # COURSES" etc: genuine SIBLING top-level categories, not real children of IT Professionals.
    # Those happened to always fail to click (likely a layout shift between when the sidebar is
    # read and when it's clicked), but relying on the click failing is luck, not a guarantee -
    # filtering out anything that's actually one of the known top-level category names means a
    # sibling category can never get mis-nested as someone else's sub-category even if a future
    # timing difference lets the click through.
    kept = [c for c in cats if c.upper() not in ALL_CATEGORIES_UPPER]
    dropped = [c for c in cats if c.upper() in ALL_CATEGORIES_UPPER]
    if dropped:
        log(f"  sidebar has {len(dropped)} sibling top-level category link(s) (cross-nav widget), "
            f"not part of this category - filtered out: {', '.join(dropped)}")
    return kept


JS_CLICK_SIDEBAR = r"""
const title = arguments[0], maxx = arguments[1];
const norm = s => (s||'').replace(/\s+/g, ' ').trim();
const as = [...document.querySelectorAll('a')].filter(a =>
  norm(a.innerText) === title && a.getBoundingClientRect().left <= maxx);
if(!as.length) return false;
as[0].scrollIntoView({block:'center'});
as[0].click();
return true;
"""

def click_sidebar_category(driver, title, attempts=4, retry_wait=1.5):
    # Retrying instead of failing after one try - the sidebar can still be mid-render (an async
    # fetch after the page itself reports "loaded") right after navigating back from a deep
    # course dive, especially once several courses/windows have been opened and closed. Trying
    # once and giving up was producing false "click fail" results that were really just "the
    # sidebar wasn't there yet".
    for attempt in range(1, attempts + 1):
        driver.execute_script("window.scrollTo(0,0);"); time.sleep(0.6)
        try:
            ok = driver.execute_script(JS_CLICK_SIDEBAR, title, SIDEBAR_MAX_X)
        except Exception:
            ok = False
        if ok:
            time.sleep(PAGE_SETTLE + 1)
            return True
        if attempt < attempts:
            time.sleep(retry_wait)
    return False


# Some course cards render a secondary "Overview" quick-peek link ahead of the real title link
# in DOM order - picking the first <a> with text longer than 2 characters can grab that instead
# of the actual course name (confirmed live: ~26% of a real category's cards mis-titled exactly
# "Overview", each also missing its Table of Contents because JS_CLICK_ASSET below was clicking
# that same wrong link into a different, incomplete page). Both scripts now skip known non-title
# labels before taking the first remaining candidate.
JS_ASSET_NAME_EXCLUDE = "['overview','about','table of contents','related items']"

JS_ASSETS = r"""
const sel = arguments[0];
const exclude = """ + JS_ASSET_NAME_EXCLUDE + r""";
return [...document.querySelectorAll(sel)].map(n => {
  const a = [...n.querySelectorAll('a')].find(x => {
    const t = (x.innerText||'').trim();
    return t.length > 2 && !exclude.includes(t.toLowerCase());
  });
  return a ? a.innerText.trim() : '';
});
"""

def asset_titles(driver):
    for sel in (ASSET_SEL, ASSET_SEL_FALLBACK):
        try:
            t = [norm(x) for x in (driver.execute_script(JS_ASSETS, sel) or [])]
            if any(t): return sel, t
        except Exception: continue
    return ASSET_SEL, []


JS_CLICK_ASSET = r"""
const sel = arguments[0], i = arguments[1];
const exclude = """ + JS_ASSET_NAME_EXCLUDE + r""";
const nodes = [...document.querySelectorAll(sel)];
if(i >= nodes.length) return false;
const a = [...nodes[i].querySelectorAll('a')].find(x => {
  const t = (x.innerText||'').trim();
  return t.length > 2 && !exclude.includes(t.toLowerCase());
});
if(!a) return false;
a.scrollIntoView({block:'center'});
a.click();
return true;
"""


def load_everything(driver, want=0):
    clicks, no_growth = 0, 0
    sel, titles = asset_titles(driver)
    last = len(titles)
    while clicks < MAX_VIEW_MORE:
        if want and last >= want: break
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1.3)
        more = None
        for label in ["view more", "show more", "load more", "see more"]:
            c = find_any(driver, [by_text("button", label), by_text("a", label),
                                  by_text("span", label)], 1.2)
            if c and c.is_displayed(): more = c; break
        if not more: break
        try: js_click(driver, more); clicks += 1
        except Exception: break
        time.sleep(2.5)
        sel, titles = asset_titles(driver)
        now = len(titles)
        log(f"    View More #{clicks} (items: {now})")
        if now <= last:
            no_growth += 1
            if no_growth >= 2: break
        else: no_growth = 0
        last = now
    driver.execute_script("window.scrollTo(0,0);"); time.sleep(0.8)
    return asset_titles(driver)


def find_next_page_button(driver, current_page_number):
    """Numbered pagination (seen as '‹ 1 2 3 ›' on some category pages) is a different
    mechanism from 'View More' - it REPLACES the visible list with a new page instead of growing
    it, so it needs its own handling rather than being folded into load_everything(). Looks for a
    clickable element whose EXACT visible text is the next page number (so '1' doesn't also match
    '12'), clicks it if found. Returns True if a next page was found and clicked."""
    target = str(current_page_number + 1)
    try:
        return bool(driver.execute_script(r"""
            const target = arguments[0];
            const els = [...document.querySelectorAll('a, button, span')];
            const el = els.find(e => (e.innerText || '').trim() === target && e.getBoundingClientRect().width > 0);
            if (!el) return false;
            const clickable = (el.tagName === 'A' || el.tagName === 'BUTTON') ? el : (el.closest('a,button') || el);
            clickable.scrollIntoView({block:'center'});
            clickable.click();
            return true;
        """, target))
    except Exception:
        return False


def open_category_list(driver, cat_url, want=0, tries=2, page_number=1):
    for t in range(tries):
        goto(driver, cat_url, PAGE_SETTLE, hard=(t > 0))
        # Restore to the same numbered page we were on before this reload, if any - a plain
        # reload always lands back on page 1.
        for p in range(1, page_number):
            if not find_next_page_button(driver, p):
                break
            time.sleep(PAGE_SETTLE)
        sel, titles = asset_titles(driver)
        if titles: break
        time.sleep(2.0)
    if page_number > 1:
        return asset_titles(driver)
    return load_everything(driver, want)


def wait_for_summary(driver, timeout=20):
    end = time.time() + timeout
    while time.time() < end:
        u = driver.current_url
        if "#summary/" in u: return u
        time.sleep(0.5)
    return ""


# ===================== COURSE DETAIL =====================

def get_panel_text(driver):
    for sel in ["div.sp-assetContainer", "div#content", "div.main-content", "body"]:
        try:
            for e in driver.find_elements(By.CSS_SELECTOR, sel):
                if e.is_displayed():
                    t = norm(e.text)
                    if len(t) > 30: return t
        except Exception: continue
    try: return norm(driver.find_element(By.TAG_NAME, "body").text)
    except Exception: return ""


LABELS = ["Course", "Expertise Level", "Status", "Duration", "Actions",
          "About", "Table of Contents", "Related Items", "Sections", "LAUNCH", "BACK"]

def field_after(text, label):
    m = re.search(rf"\b{re.escape(label)}\s*:\s*", text, re.I)
    if not m: return ""
    rest = text[m.end():]
    cut = len(rest)
    for lb in LABELS:
        if lb.lower() == label.lower(): continue
        m2 = re.search(rf"\b{re.escape(lb)}\b\s*:?", rest, re.I)
        if m2: cut = min(cut, m2.start())
    m3 = re.search(r"[\n\r]", rest)
    if m3: cut = min(cut, m3.start())
    val = norm(rest[:cut])
    val = re.sub(r"\b(LAUNCH|BACK)\b", "", val, flags=re.I)
    return norm(val)


TITLE_EXCLUDE_NAMES = {"overview", "about", "table of contents", "related items"}

def get_course_title(driver, fallback=""):
    for sel in ["h1", "h2", ".sp-assetTitle", ".course-title", ".asset-title", ".title"]:
        try:
            for e in driver.find_elements(By.CSS_SELECTOR, sel):
                text = norm(e.text)
                if e.is_displayed() and text and text.strip().lower() not in TITLE_EXCLUDE_NAMES:
                    return text
        except Exception: continue
    return fallback


def click_tab(driver, name):
    el = find_any(driver, [by_text("a", name.lower()), by_text("li", name.lower()),
                           by_text("span", name.lower()), by_text("button", name.lower()),
                           by_text("div", name.lower())], 6)
    if el:
        js_click(driver, el); time.sleep(2.0); return True
    return False


def get_about_text(driver):
    if not click_tab(driver, "About"): return ""
    for sel in [".tab-content", "#about", ".about", "[role='tabpanel']",
                ".x-tab-panel-body", ".description"]:
        try:
            for e in driver.find_elements(By.CSS_SELECTOR, sel):
                if e.is_displayed() and len(norm(e.text)) > 20: return norm(e.text)
        except Exception: continue
    return get_panel_text(driver)


def parse_about(about):
    a = clean_lines(about)
    if not a:
        return {"library_id": "", "training_credits": "", "objectives": [], "overview": ""}

    lib = ""
    m = re.search(r"\bID\s*:\s*([A-Za-z0-9_\-\.\$:]+)", a)
    if m: lib = m.group(1).strip()

    ALL = [r"\bLibrary\s*ID\b", r"\bTraining\s*Credits\b", r"\bObjectives\b",
           r"\bRelated\s*Items\b", r"\bTable\s*of\s*Contents\b", r"\bPrerequisites\b"]

    def grab(start_pat, stop_pats):
        mm = re.search(start_pat + r"\s*:?\s*", a, re.I)
        if not mm: return ""
        chunk = a[mm.end():]
        cut = len(chunk)
        for sp in stop_pats:
            m2 = re.search(sp, chunk, re.I)
            if m2: cut = min(cut, m2.start())
        return clean_lines(chunk[:cut])

    credits = grab(r"\bTraining\s*Credits\b", [p for p in ALL if "Credits" not in p])
    objectives_raw = grab(r"\bObjectives\b", [p for p in ALL if "Objectives" not in p])

    objectives = []
    for l in objectives_raw.splitlines():
        l = re.sub(r"^[\-\u2022\*\d\.\)\s]+", "", l).strip()
        if len(l) > 3: objectives.append(l)

    overview = a
    m = re.search(r"\b(Library\s*ID|Training\s*Credits|Objectives)\b", a, re.I)
    if m: overview = clean_lines(a[:m.start()])

    return {"library_id": lib, "training_credits": norm(credits),
            "objectives": objectives, "overview": overview}


def get_launch_url(driver, course_url, enabled=True):
    if not enabled: return ""
    btn = find_any(driver, [by_text("button", "launch"), by_text("a", "launch")], 6)
    if not btn: return ""
    main_h = driver.current_window_handle
    before = list(driver.window_handles)
    try:
        js_click(driver, btn); time.sleep(5.0)
        new = [h for h in driver.window_handles if h not in before]
        if new:
            driver.switch_to.window(new[0]); time.sleep(3.0)
            url = driver.current_url
            driver.close(); driver.switch_to.window(main_h); time.sleep(1.0)
            return url
        url = driver.current_url
        if url != course_url: goto(driver, course_url)
        return url
    except Exception as e:
        log(f"    launch url fail: {e}")
        try: driver.switch_to.window(main_h)
        except Exception: pass
        return ""


JS_TOC = r"""
const out = [];
const seen = new Set();
document.querySelectorAll('a').forEach(a => {
  const t = (a.innerText || '').trim();
  if (t.length < 3) return;
  const r = a.getBoundingClientRect();
  if (r.width === 0 && r.height === 0) return;
  let row = a.closest('tr') || a.closest('li');
  if (!row) {
    let p = a.parentElement, hops = 0;
    while (p && hops < 5) {
      if (/\d+\s*(Minute|Minutes|Hour|Hours)/i.test(p.innerText || '')) { row = p; break; }
      p = p.parentElement; hops++;
    }
  }
  const txt = row ? (row.innerText || '') : '';
  const dm = txt.match(/(\d+\s*(?:Minutes|Minute|Hours|Hour))/i);
  if (!dm) return;
  const key = t + '|' + dm[1];
  if (seen.has(key)) return;
  seen.add(key);
  let status = '';
  if (row) {
    const cells = row.querySelectorAll('td');
    if (cells.length >= 3) status = (cells[2].innerText || '').trim();
    if (!status) {
      const ic = row.querySelector('[title],[aria-label],img[alt]');
      if (ic) status = ic.getAttribute('title') || ic.getAttribute('aria-label') || ic.getAttribute('alt') || '';
    }
  }
  out.push({ name: t, duration: dm[1], status: status });
});
return out;
"""

TOC_EXCLUDE_NAMES = {"about", "table of contents", "related items"}

def read_toc(driver):
    try:
        rows = driver.execute_script(JS_TOC) or []
    except Exception as e:
        log(f"    toc read error: {e}")
        return []
    return [{"section_name": norm(r["name"]),
             "duration": norm(r["duration"]),
             "status": norm(r.get("status", ""))}
            for r in rows
            if norm(r.get("name")) and norm(r["name"]).strip().lower() not in TOC_EXCLUDE_NAMES]


JS_CLICK_TOC = r"""
const name = arguments[0];
const a = [...document.querySelectorAll('a')].find(x => (x.innerText||'').trim() === name);
if (!a) return false;
a.scrollIntoView({block:'center'});
a.click();
return true;
"""


def get_sections(driver, course_url, capture_urls=True):
    if not click_tab(driver, "Table of Contents"):
        log("    TOC tab not found")
        return []

    rows = []
    end = time.time() + 12
    while time.time() < end:
        rows = read_toc(driver)
        if rows: break
        time.sleep(1.0)

    if not rows:
        log("    TOC empty")
        return []

    log(f"    Sections: {len(rows)}")
    out = []
    for i, r in enumerate(rows, 1):
        sec = {"section_no": i, "section_name": r["section_name"],
               "section_url": "", "duration": r["duration"], "status": r["status"]}

        if capture_urls:
            main_h = driver.current_window_handle
            try:
                before = list(driver.window_handles)
                if driver.execute_script(JS_CLICK_TOC, r["section_name"]):
                    time.sleep(3.5)
                    new = [h for h in driver.window_handles if h not in before]
                    if new:
                        driver.switch_to.window(new[0]); time.sleep(2.5)
                        sec["section_url"] = driver.current_url
                        driver.close(); driver.switch_to.window(main_h); time.sleep(1.0)
                    else:
                        sec["section_url"] = driver.current_url
                        if driver.current_url != course_url:
                            goto(driver, course_url)
                            click_tab(driver, "Table of Contents")
                            time.sleep(1.5)
            except Exception as e:
                log(f"      section {i} url skip ({e})")
                try: driver.switch_to.window(main_h)
                except Exception: pass

        out.append(sec)
        log(f"      {i}. {r['section_name'][:55]} | {r['duration']}")
    return out


def scrape_open_course(driver, cat_path, list_title, capture_launch, capture_sections,
                       main_category_url=None, sub_category_url=None):
    url = driver.current_url
    panel = get_panel_text(driver)
    # The course-list page (before ever clicking in) already has the full, correct,
    # untruncated title - no need to hunt through the detail page's headings for it, which is
    # exactly what was misfiring (the About tab's own "Overview" heading was sometimes matched
    # instead of the real title). Only fall back to scanning the detail page if the list somehow
    # didn't give us a title at all.
    title = norm(list_title) or get_course_title(driver)

    duration = field_after(panel, "Course")
    if not duration:
        m = re.search(r"(\d+\s*(?:Minutes|Minute|Hours|Hour))", panel, re.I)
        duration = norm(m.group(1)) if m else ""

    about = get_about_text(driver)
    ab = parse_about(about)

    launch = get_launch_url(driver, url, capture_launch)
    goto(driver, url, 2.0)
    sections = get_sections(driver, url, capture_sections)

    return {
        "category_path": cat_path,
        "main_category_url": main_category_url,
        "sub_category_url": sub_category_url,
        "course_url": url,
        "name": title,
        "course_duration": duration,
        "expertise_level": field_after(panel, "Expertise Level"),
        "status": field_after(panel, "Status"),
        "library_id": ab["library_id"],
        "training_credits": ab["training_credits"],
        "objectives": ab["objectives"],
        "about_overview": ab["overview"],
        "about_full": clean_lines(about),
        "launch_url": launch,
        "sections_count": len(sections),
        "section_names": [s["section_name"] for s in sections],
        "sections": sections,
        "scraped_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }


# ===================== LIST CRAWLING =====================

def scrape_current_list(driver, cat_url, cat_path, data, done_urls, prog, ppath, dpath,
                        limit, scraped, capture_launch, capture_sections,
                        main_category_url=None, sub_category_url=None, page_number=1):
    if page_number == 1:
        sel, titles = load_everything(driver)
    else:
        # Already navigated to this exact page by the caller (either freshly, via
        # find_next_page_button, or restored to it by open_category_list after a reload) -
        # numbered pages don't have their own "View More" on top of pagination.
        sel, titles = asset_titles(driver)
    total = len(titles)
    log(f"  {total} courses found" + (f" (page {page_number})" if page_number > 1 else ""))
    if not total:
        return scraped, True

    # Fast pre-check by title, before ever clicking - lets an already-fully-scraped page skip
    # in a fraction of a second per course instead of clicking into every one just to discover
    # it's already there (which also costs a full page-reload afterward). The per-URL resume
    # position below doesn't survive across separate script runs (this listing's own URL isn't
    # stable between visits), so this title check is what actually makes re-running fast.
    # Known, accepted tradeoff: if two different courses in the same list ever share the exact
    # same title, the second would be wrongly treated as already-done - course titles are
    # effectively unique in this catalog, so this is safe in practice.
    done_titles = {norm(d.get("name")).lower() for d in data if d.get("name")}

    # Numbered pages share the same cat_url but need their own progress-tracking slot - otherwise
    # page 2's positions would collide with page 1's.
    page_key = cat_url if page_number == 1 else f"{cat_url}#page{page_number}"
    start = prog.get(page_key, 0)
    if start:
        log(f"  {start} already done, continuing from there")

    for i in range(start, total):
        if limit and scraped >= limit:
            log("Limit reached")
            save_json(ppath, prog)
            return scraped, False

        list_title = titles[i] if i < len(titles) else ""

        if norm(list_title).lower() in done_titles:
            log(f"  [{i+1}/{total}] {list_title[:65]} - already done (title match), skip")
            prog[page_key] = i + 1
            continue

        log(f"  [{i+1}/{total}] {list_title[:65]}")
        try:
            if not driver.execute_script(JS_CLICK_ASSET, sel, i):
                log("    click failed"); continue
            url = wait_for_summary(driver)
            if not url:
                log("    summary page did not open")
                sel, titles = open_category_list(driver, cat_url, total, page_number=page_number)
                continue
            if url in done_urls:
                log("    already done, skip")
            else:
                rec = scrape_open_course(driver, cat_path, list_title,
                                         capture_launch, capture_sections,
                                         main_category_url, sub_category_url)
                data.append(rec); done_urls.add(url); done_titles.add(norm(rec.get("name")).lower()); scraped += 1
                save_json(dpath, data)
                log(f"    OK: {rec['name'][:50]} | {rec['course_duration']} | "
                    f"{rec['sections_count']} sections")
        except Exception as e:
            log(f"    FAIL: {e}"); shot(driver, "course_fail")
        finally:
            prog[page_key] = i + 1
            save_json(ppath, prog)
            sel, titles = open_category_list(driver, cat_url, total, page_number=page_number)
            total = max(total, len(titles))

    # This page is done - check for numbered pagination (a different mechanism from "View More",
    # which was already handled by load_everything() above for page 1). If there's a next page,
    # recurse into it as its own progress-tracked unit sharing the same data/done_urls/prog.
    if find_next_page_button(driver, page_number):
        time.sleep(PAGE_SETTLE + 1)
        return scrape_current_list(
            driver, cat_url, cat_path, data, done_urls, prog, ppath, dpath,
            limit, scraped, capture_launch, capture_sections,
            main_category_url, sub_category_url, page_number=page_number + 1)

    return scraped, True


MAX_SIDEBAR_DEPTH = 6  # safety cap - confirmed real nesting has gone 3 levels deep
                       # (COURSES FOR IT PROFESSIONALS > Data Science > Data Analyst to Data
                       # Scientist > Data Science Track 4: Data Scientist), extra margin kept
                       # above that in case some other branch nests one level deeper still.


def crawl_sidebar_tree(driver, parent_url, category_path, data, done_urls, prog, ppath, dpath,
                       limit, scraped, capture_launch, capture_sections, main_category_url,
                       depth=1, seen_category_sets=None, skipped=None):
    """Recursively walks a category's sidebar sub-categories to whatever depth they actually go -
    some categories nest two or three levels deep (e.g. Microsoft/Desktop Courses > Microsoft 365
    > Excel 2026), not just one. Scrapes each sub-category's own direct course list, then checks
    whether IT also has its own sidebar and descends into that too, stopping only once a level has
    no further sub-categories of its own."""
    if seen_category_sets is None:
        seen_category_sets = []
    if skipped is None:
        skipped = []

    if depth > MAX_SIDEBAR_DEPTH:
        log(f"  Reached max sidebar depth ({MAX_SIDEBAR_DEPTH}), stopping: {category_path}")
        skipped.append(f"{category_path} (max depth {MAX_SIDEBAR_DEPTH} reached - may have unscraped children)")
        return scraped, True

    hard_reload(driver, parent_url)
    cats = get_sidebar_categories(driver)
    if not cats:
        return scraped, True

    # Cycle guard: some sidebars are a static/global "browse by category" panel that shows the
    # SAME set of names on every page rather than genuine children of the current node - without
    # this, recursion explores an exponentially growing number of meaningless combinations of the
    # same handful of categories (seen live: 11 categories repeating identically 3-4 levels deep,
    # 1000+ dead-end paths, over an hour wasted, zero new courses - the course counts were
    # byte-for-byte identical at every depth, confirming it was the same static content, not a
    # real deepening tree). If this exact set of names already appeared earlier in this same
    # descent, it's a repeat - stop here instead of re-exploring it.
    cat_set = frozenset(c.strip().lower() for c in cats)
    if cat_set in seen_category_sets:
        log(f"  '{category_path}''s sidebar already appeared earlier on this path - looks like a static/repeating list, stopping here")
        return scraped, True

    log(f"  '{category_path}' has {len(cats)} sub-categories")
    for ci, cat in enumerate(cats, 1):
        log(f"=== [{ci}/{len(cats)}] Sub category: {category_path} > {cat} ===")
        hard_reload(driver, parent_url)
        if not click_sidebar_category(driver, cat):
            log("  click failed, skip")
            skipped.append(f"{category_path} > {cat} (click failed)")
            continue
        cat_url = driver.current_url

        # Safety: only ever descend into genuine course-browsing pages (#browse/<id> URLs, the
        # pattern every real category/sub-category page on this site has used so far). The
        # generic "find sidebar links on the left of the screen" heuristic has no way to tell a
        # course category apart from some other menu that happens to render in the same screen
        # region (e.g. an admin/user-management panel, visible because this account is an admin
        # account) - without this check it can wander arbitrarily deep into unrelated, sensitive
        # parts of the site. Bail out the moment we land somewhere that isn't course browsing.
        if "#browse/" not in cat_url:
            log(f"  '{cat}' is not a course-browsing page ({cat_url}) - bailing out for safety, skip")
            skipped.append(f"{category_path} > {cat} (not a course-browsing page - safety exclusion, likely a non-catalog menu)")
            continue

        new_path = f"{category_path} > {cat}"

        scraped, finished = scrape_current_list(
            driver, cat_url, new_path, data, done_urls, prog, ppath, dpath,
            limit, scraped, capture_launch, capture_sections,
            main_category_url=main_category_url, sub_category_url=cat_url)
        if not finished:
            return scraped, False

        scraped, finished = crawl_sidebar_tree(
            driver, cat_url, new_path, data, done_urls, prog, ppath, dpath,
            limit, scraped, capture_launch, capture_sections, main_category_url,
            depth=depth + 1, seen_category_sets=seen_category_sets + [cat_set], skipped=skipped)
        if not finished:
            return scraped, False

    return scraped, True


def log_skip_report(top_category, skipped):
    """Printed once at the end of every category's crawl (successful or limit-cut-short) so a run
    is verifiable after the fact instead of requiring someone to scroll the whole live log looking
    for scattered 'click fail' lines - the exact concern raised after the deep-nesting incident:
    confirm nothing under a category was silently missed."""
    if not skipped:
        log(f"[SKIP-REPORT] category={top_category} skipped=0 (nothing was skipped)")
        return
    log(f"[SKIP-REPORT] category={top_category} skipped={len(skipped)}:")
    for item in skipped:
        log(f"  - {item}")


def crawl(driver, top_category, mode="top", limit=None,
          capture_launch=True, capture_sections=True, tag=None):
    tag = tag or file_tag(mode, top_category)
    dpath = OUT_DIR / f"courses_data_{tag}.json"
    ppath = OUT_DIR / f"progress_{tag}.json"
    data = load_json(dpath, [])
    prog = load_json(ppath, {})
    done_urls = {d.get("course_url") for d in data if d.get("course_url")}
    scraped = 0

    open_library(driver)
    click_top_category(driver, top_category)
    top_url = driver.current_url
    log(f"Top page url: {top_url}")

    # ---- STEP 1: the main page's own list ----
    if mode in ("top", "both"):
        log(f"=== MAIN LIST: {top_category} ===")
        scraped, finished = scrape_current_list(
            driver, top_url, top_category, data, done_urls, prog, ppath, dpath,
            limit, scraped, capture_launch, capture_sections,
            main_category_url=top_url, sub_category_url=None)
        if not finished:
            return data
        log(f"MAIN LIST complete. {len(data)} courses saved so far")

    # ---- STEP 2: sidebar sub categories, recursively (some nest 2-3 levels deep) ----
    skipped_categories = []
    if mode in ("sidebar", "both"):
        scraped, finished = crawl_sidebar_tree(
            driver, top_url, top_category, data, done_urls, prog, ppath, dpath,
            limit, scraped, capture_launch, capture_sections, main_category_url=top_url,
            skipped=skipped_categories)
        if not finished:
            log_skip_report(top_category, skipped_categories)
            return data

    log_skip_report(top_category, skipped_categories)
    return data


# ===================== DEBUG =====================

def dump_debug(driver):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    lines = [f"URL: {driver.current_url}", f"TITLE: {driver.title}"]
    lines.append("\n--- SIDEBAR CATEGORIES ---")
    lines += get_sidebar_categories(driver)
    sel, titles = asset_titles(driver)
    lines.append(f"\n--- ASSETS (selector {sel}) : {len(titles)} ---")
    lines += titles
    try:
        html = driver.execute_script(
            "const n=document.querySelector(arguments[0]); return n ? n.outerHTML : 'NONE';", sel)
        lines.append("\n--- FIRST ASSET'S HTML ---"); lines.append(html[:6000])
    except Exception as e:
        lines.append(f"outerHTML error: {e}")
    (OUT_DIR / "debug_dom.txt").write_text("\n".join(map(str, lines)), encoding="utf-8")
    shot(driver, "list_page")
    log("Debug file: output/debug_dom.txt")


# ===================== EXPORT =====================

def export_notepad(data, path):
    out = []
    for i, r in enumerate(data, 1):
        out.append("=" * 90)
        out.append(f"{i}. {r.get('name','')}")
        out.append("=" * 90)
        out.append(f"Category Path   : {r.get('category_path','')}")
        out.append(f"Course          : {r.get('course_duration','')}")
        out.append(f"Expertise Level : {r.get('expertise_level','')}")
        out.append(f"Status          : {r.get('status','')}")
        out.append(f"Library ID      : {r.get('library_id','')}")
        out.append(f"Training Credits: {r.get('training_credits','')}")
        out.append(f"Course URL      : {r.get('course_url','')}")
        out.append(f"Launch URL      : {r.get('launch_url','')}")
        if r.get("about_overview"):
            out.append("\n--- ABOUT ---")
            out.append(r["about_overview"])
        if r.get("objectives"):
            out.append("\n--- OBJECTIVES ---")
            out += [f"  {n+1}. {o}" for n, o in enumerate(r["objectives"])]
        secs = r.get("sections", [])
        out.append(f"\n--- TABLE OF CONTENTS ({len(secs)} sections) ---")
        for s in secs:
            out.append(f"  {s.get('section_no','')}. {s.get('section_name','')}"
                       f"   [{s.get('duration','')}]  status: {s.get('status','')}")
            if s.get("section_url"):
                out.append(f"      {s['section_url']}")
        out.append("")
    Path(path).write_text("\n".join(out), encoding="utf-8")
    log(f"Notepad file created: {path}")


def export_excel(data, path, notepad_path=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("Run: pip install openpyxl"); return

    wb = Workbook()

    # ---------- Sheet 1: Courses ----------
    ws = wb.active; ws.title = "Courses"
    ws.append(["Category Path", "Course Name", "Course Duration", "Expertise Level",
               "Status", "Library ID", "Training Credits", "Course URL", "Launch URL",
               "Sections Count", "TOC Section Names", "TOC Section Durations",
               "Objectives", "About"])

    for r in data:
        sec_objs = r.get("sections", [])
        if sec_objs:
            names = [s.get("section_name", "") for s in sec_objs if s.get("section_name")]
            durs  = [s.get("duration", "") for s in sec_objs if s.get("section_name")]
        else:
            names = [s for s in (r.get("section_names") or []) if s]
            durs  = [""] * len(names)

        ws.append([
            r.get("category_path", ""),
            r.get("name", ""),
            r.get("course_duration", ""),
            r.get("expertise_level", ""),
            r.get("status", ""),
            r.get("library_id", ""),
            r.get("training_credits", ""),
            r.get("course_url", ""),
            r.get("launch_url", ""),
            r.get("sections_count", len(names)),
            ", ".join(names)[:32000],
            ", ".join(durs)[:32000],
            ", ".join(r.get("objectives", []))[:32000],
            (r.get("about_full") or r.get("about", ""))[:32000],
        ])

    # ---------- Sheet 2: Sections ----------
    ws2 = wb.create_sheet("Sections")
    ws2.append(["Category Path", "Course Name", "Library ID", "Course URL",
                "Section No", "Section Name", "Section Duration", "Section URL", "Status"])
    for r in data:
        for s in r.get("sections", []):
            ws2.append([r.get("category_path", ""), r.get("name", ""),
                        r.get("library_id", ""), r.get("course_url", ""),
                        s.get("section_no", ""), s.get("section_name", ""),
                        s.get("duration", ""), s.get("section_url", ""), s.get("status", "")])

    fill = PatternFill("solid", start_color="1F2937")
    wrap_cols = {"Courses": [11, 12, 13, 14], "Sections": []}
    for sheet, widths in ((ws, [26, 44, 15, 15, 13, 22, 16, 55, 55, 13, 70, 30, 60, 60]),
                          (ws2, [26, 40, 22, 55, 10, 50, 16, 55, 13])):
        for c in sheet[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
            c.alignment = Alignment(vertical="center")
        for i, w in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = w
        for row in sheet.iter_rows(min_row=2):
            for c in row:
                if c.column in wrap_cols.get(sheet.title, []):
                    c.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    wb.save(path)
    log(f"Excel file created: {path}")
    export_notepad(data, notepad_path or (OUT_DIR / "courses_readable.txt"))


# ===================== SQL EXPORT =====================
# Column limits below mirror the EF Core configs exactly (CourseConfiguration.cs,
# MainCourseCategoryConfiguration.cs, SubCourseCategoryConfiguration.cs,
# CourseSectionConfiguration.cs) - keep in sync if those ever change.

def sql_str(value, max_len=None):
    """NULL / N'escaped' literal. Truncates the RAW value first, THEN doubles quotes -
    doing it the other way round can cut a doubled '' pair in half at the boundary and
    corrupt the literal. Truncating first is also correct because SQL Server's
    nvarchar(n) limit counts the stored (unescaped) character length, not the literal's
    escaped text length."""
    if value is None:
        return "NULL"
    s = str(value).strip()
    if not s:
        return "NULL"
    if max_len is not None and len(s) > max_len:
        s = s[:max_len]
    return "N'" + s.replace("'", "''") + "'"


def sql_datetime(value):
    return f"'{value}'" if value else "SYSUTCDATETIME()"


def export_sql(data, path):
    if not data:
        log("SQL export skipped: no data")
        return

    main_categories = {}   # name -> {"type_id": int, "url": str|None}
    sub_categories = {}    # (main_name, sub_name) -> {"url": str|None}

    def split_path(cat_path):
        parts = [p.strip() for p in (cat_path or "").split(">")]
        main_name = parts[0] if parts and parts[0] else ""
        sub_name = parts[1] if len(parts) > 1 and parts[1] else None
        return main_name, sub_name

    for r in data:
        main_name, sub_name = split_path(r.get("category_path"))
        if not main_name:
            continue

        entry = main_categories.setdefault(
            main_name, {"type_id": type_id_for_category(main_name), "url": None})
        if not entry["url"] and r.get("main_category_url"):
            entry["url"] = r["main_category_url"]

        if sub_name:
            key = (main_name, sub_name)
            sub_entry = sub_categories.setdefault(key, {"url": None})
            if not sub_entry["url"] and r.get("sub_category_url"):
                sub_entry["url"] = r["sub_category_url"]

    lines = [
        "-- Auto-generated by skillport_scraper.py export_sql() - review before running.",
        f"-- Generated at: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"-- Courses in this file: {len(data)}",
        "SET NOCOUNT ON;",
        "GO",
        "",
    ]

    for name, info in main_categories.items():
        type_id = info["type_id"]
        lines += [
            f"-- Main category: {name}",
            f"IF NOT EXISTS (SELECT 1 FROM MainCourseCategories WHERE CategoryName = {sql_str(name, 255)} AND TypeID = {type_id})",
            "BEGIN",
            "    INSERT INTO MainCourseCategories (TypeID, CategoryName, CategoryURL, DisplayOrder, IsActive, CreatedAt)",
            f"    VALUES ({type_id}, {sql_str(name, 255)}, {sql_str(info['url'], 2000)}, "
            f"(SELECT ISNULL(MAX(DisplayOrder), 0) + 1 FROM MainCourseCategories WHERE TypeID = {type_id}), 1, SYSUTCDATETIME());",
            "END",
            "GO",
            "",
        ]

    for (main_name, sub_name), info in sub_categories.items():
        type_id = main_categories[main_name]["type_id"]
        lines += [
            f"-- Sub category: {main_name} > {sub_name}",
            "DECLARE @ParentCategoryID INT = "
            f"(SELECT CategoryID FROM MainCourseCategories WHERE CategoryName = {sql_str(main_name, 255)} AND TypeID = {type_id});",
            f"IF NOT EXISTS (SELECT 1 FROM SubCourseCategories WHERE SubCategoryName = {sql_str(sub_name, 500)} AND CategoryID = @ParentCategoryID)",
            "BEGIN",
            "    INSERT INTO SubCourseCategories (CategoryID, SubCategoryName, SubCategoryURL, DisplayOrder, IsActive, CreatedAt)",
            f"    VALUES (@ParentCategoryID, {sql_str(sub_name, 500)}, {sql_str(info['url'], 2000)}, "
            "(SELECT ISNULL(MAX(DisplayOrder), 0) + 1 FROM SubCourseCategories WHERE CategoryID = @ParentCategoryID), 1, SYSUTCDATETIME());",
            "END",
            "GO",
            "",
        ]

    skipped = 0
    for r in data:
        main_name, sub_name = split_path(r.get("category_path"))
        course_url = (r.get("course_url") or "").strip()
        if not main_name or not course_url:
            skipped += 1
            continue
        type_id = main_categories[main_name]["type_id"]

        # AboutContent is deliberately left NULL, not about_full - about_full re-merges the
        # overview paragraph with the Library ID and the full Objectives list into one blob,
        # which would duplicate data that already has its own column (SkillsoftCourseCode,
        # Objectives). Only the clean, already-separated overview paragraph goes into
        # OverviewContent.
        overview = r.get("about_overview") or ""
        objectives_text = "\n".join(o.strip() for o in (r.get("objectives") or []) if o.strip())

        lines.append(f"-- Course: {r.get('name','')}")
        lines.append(
            f"DECLARE @CategoryID INT = (SELECT CategoryID FROM MainCourseCategories WHERE CategoryName = {sql_str(main_name, 255)} AND TypeID = {type_id});")
        if sub_name:
            lines.append(
                f"DECLARE @SubCategoryID INT = (SELECT SubCategoryID FROM SubCourseCategories WHERE SubCategoryName = {sql_str(sub_name, 500)} AND CategoryID = @CategoryID);")
        else:
            lines.append("DECLARE @SubCategoryID INT = NULL;")

        lines.append(f"IF EXISTS (SELECT 1 FROM Courses WHERE CourseURL = {sql_str(course_url, 2000)})")
        lines.append("BEGIN")
        lines.append("    UPDATE Courses SET")
        lines.append("        CategoryID = @CategoryID,")
        lines.append("        SubCategoryID = @SubCategoryID,")
        lines.append(f"        CourseTitle = {sql_str(r.get('name'), 1000)},")
        lines.append(f"        LaunchURL = {sql_str(r.get('launch_url'))},")
        lines.append(f"        Duration = {sql_str(r.get('course_duration'), 100)},")
        lines.append(f"        ExpertiseLevel = {sql_str(r.get('expertise_level'), 255)},")
        lines.append(f"        Status = {sql_str(r.get('status'), 255)},")
        lines.append("        AboutContent = NULL,")
        lines.append(f"        OverviewContent = {sql_str(overview)},")
        lines.append(f"        SkillsoftCourseCode = {sql_str(r.get('library_id'), 500)},")
        lines.append(f"        ScrapedAt = {sql_datetime(r.get('scraped_at'))},")
        lines.append(f"        CourseSourceUrl = {sql_str(course_url, 1000)},")
        lines.append("        SourceStatus = N'Scraped',")
        lines.append(f"        Objectives = {sql_str(objectives_text)}")
        lines.append(f"    WHERE CourseURL = {sql_str(course_url, 2000)};")
        lines.append("END")
        lines.append("ELSE")
        lines.append("BEGIN")
        lines.append("    INSERT INTO Courses (CategoryID, SubCategoryID, CourseTitle, CourseURL, LaunchURL, Duration,")
        lines.append("        ExpertiseLevel, Status, AboutContent, OverviewContent, ImageURL, SkillsoftCourseCode,")
        lines.append("        DisplayOrder, IsActive, ScrapedAt, CourseSourceUrl, SourceStatus, Objectives)")
        lines.append(
            "    VALUES (@CategoryID, @SubCategoryID, "
            f"{sql_str(r.get('name'), 1000)}, {sql_str(course_url, 2000)}, {sql_str(r.get('launch_url'))}, "
            f"{sql_str(r.get('course_duration'), 100)}, {sql_str(r.get('expertise_level'), 255)}, "
            f"{sql_str(r.get('status'), 255)}, NULL, {sql_str(overview)}, NULL, "
            f"{sql_str(r.get('library_id'), 500)}, "
            "(SELECT ISNULL(MAX(DisplayOrder), 0) + 1 FROM Courses WHERE CategoryID = @CategoryID), 1, "
            f"{sql_datetime(r.get('scraped_at'))}, {sql_str(course_url, 1000)}, N'Scraped', {sql_str(objectives_text)});")
        lines.append("END")

        # Sections: full replace-on-rescrape (delete + reinsert) rather than diffing the TOC -
        # same batch as the upsert above, so @CourseId resolves whether the row was just
        # inserted or already existed.
        lines.append(f"DECLARE @CourseId BIGINT = (SELECT CourseID FROM Courses WHERE CourseURL = {sql_str(course_url, 2000)});")
        lines.append("DELETE FROM CourseSections WHERE CourseID = @CourseId;")
        sections = r.get("sections") or []
        if sections:
            lines.append("INSERT INTO CourseSections (CourseID, SectionName, SectionURL, Duration, Status, DisplayOrder, CreatedAt)")
            values = [
                "    (@CourseId, "
                f"{sql_str(s.get('section_name'), 1000)}, {sql_str(s.get('section_url'), 2000)}, "
                f"{sql_str(s.get('duration'), 100)}, {sql_str(s.get('status'), 255)}, "
                f"{s.get('section_no', 0)}, SYSUTCDATETIME())"
                for s in sections
            ]
            lines.append("VALUES\n" + ",\n".join(values) + ";")
        lines.append("GO")
        lines.append("")

    if skipped:
        log(f"SQL export: {skipped} course(s) skipped (no category or no course_url)")

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text("\n".join(lines), encoding="utf-8")
    log(f"[SQL] Exported SQL to: {path}")


# ===================== ENTRY =====================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--category", action="append", default=None,
                    help="a single category name, ALL for every top category, or pass "
                         "--category multiple times to choose several categories (e.g. --category \"PMI COURSES\" "
                         "--category \"SIX SIGMA COURSES\")")
    ap.add_argument("--mode", choices=["top", "sidebar", "both"], default="top")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--headless", action="store_true")
    ap.add_argument("--debug", action="store_true")
    ap.add_argument("--no-launch", action="store_true")
    ap.add_argument("--no-section-urls", action="store_true")
    ap.add_argument("--reset", action="store_true")
    ap.add_argument("--export-only", action="store_true")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # --category can now be passed multiple times (argparse action="append") to run several
    # specific categories in one go, not just one or ALL - e.g.
    # --category "SIX SIGMA COURSES" --category "COMPLIANCE COURSES".
    if not args.category:
        cats = [DEFAULT_CATEGORY]
    elif any(c.strip().upper() == "ALL" for c in args.category):
        cats = ALL_CATEGORIES
    else:
        cats = args.category

    def paths_for(cat):
        t = file_tag(args.mode, cat)
        return (OUT_DIR / f"courses_data_{t}.json",
                OUT_DIR / f"progress_{t}.json",
                OUT_DIR / f"skillport_courses_{t}.xlsx",
                OUT_DIR / f"courses_readable_{t}.txt",
                OUT_DIR / f"courses_insert_{t}.sql", t)

    if args.export_only:
        combined = []
        for cat in cats:
            dp, _, xp, tp, sp, _ = paths_for(cat)
            d = load_json(dp, [])
            if d:
                export_excel(d, xp, tp); export_sql(d, sp); combined += d
            else:
                log(f"{dp.name} is empty")
        if len(cats) > 1 and combined:
            export_excel(combined, OUT_DIR / "skillport_ALL_combined.xlsx",
                         OUT_DIR / "courses_readable_ALL_combined.txt")
            export_sql(combined, OUT_DIR / "courses_insert_ALL_combined.sql")
            log(f"Combined file created. Total courses: {len(combined)}")
        return

    if args.reset:
        for cat in cats:
            dp, pp, _, _, _, t = paths_for(cat)
            for p in (pp, dp):
                if p.exists(): p.unlink()
            log(f"Progress reset ({t})")

    driver = make_driver(args.headless)
    combined = []
    try:
        login(driver)

        if args.debug:
            open_library(driver)
            click_top_category(driver, cats[0])
            log(f"Sidebar categories: {get_sidebar_categories(driver)}")
            load_everything(driver)
            dump_debug(driver)
            input("Debug done. Check the browser, then press Enter...")
            return

        for ci, cat in enumerate(cats, 1):
            dp, pp, xp, tp, sp, t = paths_for(cat)
            log("#" * 70)
            log(f"### TOP CATEGORY [{ci}/{len(cats)}]: {cat}")
            log("#" * 70)
            try:
                data = crawl(driver, cat, args.mode, args.limit,
                             not args.no_launch, not args.no_section_urls, tag=t)
            except Exception as e:
                log(f"CATEGORY FAIL ({cat}): {e}")
                try: shot(driver, "cat_fail")
                except Exception: pass
                data = load_json(dp, [])

                # A category-level failure (as opposed to a single-course failure, which is
                # already handled inside scrape_current_list) usually means the browser session
                # itself died (crashed Chrome, dropped WebDriver connection, etc.) - reusing a
                # dead session for the remaining categories just fails every one of them
                # instantly with zero real attempts, silently losing whatever's left of a long
                # ALL-categories run. Recover by quitting and relaunching a fresh session and
                # logging back in before moving on, so one crash doesn't take out the rest of
                # the run.
                log("Category failed - recovering the browser session (new session + re-login)...")
                force_kill_driver(driver)
                recovered = False
                for attempt in range(1, 3):
                    try:
                        driver = make_driver(args.headless)
                        login(driver)
                        log("Session recovered, continuing with the next category")
                        recovered = True
                        break
                    except Exception as e2:
                        log(f"Session recovery attempt {attempt}/2 failed: {e2}")
                        force_kill_driver(driver)
                        time.sleep(5)
                if not recovered:
                    log("Session recovery failed completely - will still try the next category")
            if data:
                export_excel(data, xp, tp)
                export_sql(data, sp)
                combined += data
            log(f"### {cat} done. Courses: {len(data)}")
            log(f"[SUMMARY] category={cat} type_id={type_id_for_category(cat)} courses={len(data)}")

    except KeyboardInterrupt:
        log("Stopped. Progress is saved - run again to continue from here.")
    except Exception as e:
        log(f"FATAL: {e}"); traceback.print_exc(); shot(driver, "fatal")
    finally:
        try: driver.quit()
        except Exception: pass

    if not combined:
        for cat in cats:
            dp, _, _, _, _, _ = paths_for(cat)
            combined += load_json(dp, [])
    if combined and len(cats) > 1:
        export_excel(combined, OUT_DIR / "skillport_ALL_combined.xlsx",
                     OUT_DIR / "courses_readable_ALL_combined.txt")
        export_sql(combined, OUT_DIR / "courses_insert_ALL_combined.sql")
        log(f"Combined file created. Total courses: {len(combined)}")


if __name__ == "__main__":
    main()